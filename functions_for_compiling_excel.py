import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.utils import get_column_letter
import pandas as pd

# this generates .txt files by stripping extra lines with rmsd_text_to_column
def process_files_rmsd(folder_path, file_counter, F_or_G, mutant, target_chain):
    delimiter = ' '
    if F_or_G == 'Factin':
        structure_type = [f'{mutant}', 'WT']
        file_naming = []
        subdomain_naming = {}
        for structure in structure_type:
            naming_phrases = [
                f'{structure}_Chain_{target_chain}_protein',
                f'{structure}_Chain_{target_chain}_SD'
            ]
            subdomain_numbers = {
                f'{structure}_Chain_{target_chain}_SD': 1
            }
            file_naming.extend(naming_phrases)
            subdomain_naming.update(subdomain_numbers)

        reset_counter = False

        for naming_phrase in file_naming:
            for filename in os.listdir(folder_path):
                if naming_phrase in filename:
                    input_file_path = os.path.join(folder_path, filename)
                    if reset_counter:
                        file_counter = 1
                        reset_counter = False

                    if naming_phrase in subdomain_naming:
                        output_file_path = os.path.join(folder_path,
                                                        f"{file_counter}_{naming_phrase}_RMSD_SD{subdomain_naming[naming_phrase]}.txt")
                        subdomain_naming[naming_phrase] += 1
                    else:
                        output_file_path = os.path.join(folder_path, f"{file_counter}_{naming_phrase}.txt")

                    # for any of the .xvg gromacs files, strip any lines that don't have data (aka starts with #, @, or &)
                    # this is called in the process_files_rmsd below
                    with open(input_file_path, 'r') as f:
                        lines = f.readlines()
                    # Remove lines starting with '#' or '@' (common in .xvg files)
                    lines = [line.strip() for line in lines if
                             not line.startswith('#') and not line.startswith('@') and not line.startswith(
                                 '&') and line.strip()]
                    two_columns = []
                    for line in lines:
                        split_line = line.split(delimiter)
                        two_columns.extend([(split_line[0], ' '.join(split_line[1:]))])
                    with open(output_file_path, 'w') as f_out:
                        for row in two_columns:
                            f_out.write('\t'.join(row) + '\n')
                    file_counter += 1
    if F_or_G == 'Gactin':
        structure_type = [f'{mutant}', 'WT']
        file_naming = []
        subdomain_naming = {}
        for structure in structure_type:
            naming_phrases = [
                f'{structure}_Chain_{target_chain}_protein',
                f'{structure}_Chain_{target_chain}_SD'
            ]
            subdomain_numbers = {
                f'{structure}_Chain_{target_chain}_SD': 1
            }
            file_naming.extend(naming_phrases)
            subdomain_naming.update(subdomain_numbers)

        reset_counter = False
        # first we output a .txt file with a new name from the .xvg file
        for naming_phrase in file_naming:
            for filename in os.listdir(folder_path):
                if naming_phrase in filename:
                    input_file_path = os.path.join(folder_path, filename)
                    if reset_counter:
                        file_counter = 1
                        reset_counter = False

                    if naming_phrase in subdomain_naming:
                        output_file_path = os.path.join(folder_path,
                                                        f"{file_counter}_{naming_phrase}_RMSD_SD{subdomain_naming[naming_phrase]}.txt")
                        subdomain_naming[naming_phrase] += 1
                    else:
                        output_file_path = os.path.join(folder_path, f"{file_counter}_{naming_phrase}.txt")

                    # for any of the .xvg gromacs files, strip any lines that don't have data (aka starts with #, @, or &)
                    # this is called in the process_files_rmsd below
                    with open(input_file_path, 'r') as f:
                        lines = f.readlines()
                    # Remove lines starting with '#' or '@' (common in .xvg files)
                    lines = [line.strip() for line in lines if
                             not line.startswith('#') and not line.startswith('@') and not line.startswith(
                                 '&') and line.strip()]
                    two_columns = []
                    for line in lines:
                        split_line = line.split(delimiter)
                        two_columns.extend([(split_line[0], ' '.join(split_line[1:]))])
                    with open(output_file_path, 'w') as f_out:
                        for row in two_columns:
                            f_out.write('\t'.join(row) + '\n')
                    file_counter += 1
    # then for all of the .txt files generated, open them, convert nm to A, save as 'processed_'.txt
    for filename in os.listdir(folder_path):
        if filename.endswith('.txt'):  # Process only .txt files
            input_file_path = os.path.join(folder_path, filename)
            output_file_path = os.path.join(folder_path, f"processed_{filename}")

            # Read the contents of the .txt file
            with open(input_file_path, 'r') as txt_file:
                lines = txt_file.readlines()

            # Apply formula to multiply values by 10 (if the value is not 0)
            modified_lines = []
            for line in lines:
                line_data = line.strip().split('\t')  # Assuming tab-separated data, modify delimiter if needed
                if len(line_data) >= 2:
                    try:
                        value = float(line_data[1])
                        if value != 0:
                            line_data.append(str(value * 10))
                        else:
                            line_data.append('0')
                    except ValueError:
                        line_data.append('Invalid')  # Or handle the invalid value case differently

                modified_lines.append('\t'.join(line_data))

            # Write the modified data to a new output file
            with open(output_file_path, 'w') as output_txt_file:
                output_txt_file.write('\n'.join(modified_lines))


# this function generates an excel file for compiled data and makes/fills in the RMSD sheets
def make_compiled_excel(F_or_G, mutant, output_path, number_of_timepoints, target_chain):
    # Create a new workbook for compiling the data
    wb = Workbook()
    if F_or_G == 'Factin':
        ws1 = wb.active
        ws1.title = f'{mutant}_RMSD'
        rmsd_path1 = f'{output_path}Replicate1/RMSD/'
        rmsd_path2 = f'{output_path}Replicate2/RMSD/'
        rmsd_path3 = f'{output_path}Replicate3/RMSD/'
        copy_time_data_rmsd(rmsd_path1, F_or_G, mutant, ws1, number_of_timepoints, target_chain)
        # add title information for RMSD
        ws1['B2'] = 'Time'
        ws1.merge_cells('C1:E1')
        ws1['C1'] = f'Chain{target_chain}'
        ws1['C2'] = 'Replicate1'
        ws1['D2'] = 'Replicate2'
        ws1['E2'] = 'Replicate3'
        ws1.merge_cells('F1:H1')
        ws1['F1'] = f'Chain{target_chain} SD1'
        ws1['F2'] = 'Replicate1'
        ws1['G2'] = 'Replicate2'
        ws1['H2'] = 'Replicate3'
        ws1.merge_cells('I1:K1')
        ws1['I1'] = f'Chain{target_chain} SD2'
        ws1['I2'] = 'Replicate1'
        ws1['J2'] = 'Replicate2'
        ws1['K2'] = 'Replicate3'
        ws1.merge_cells('L1:N1')
        ws1['L1'] = f'Chain{target_chain} SD3'
        ws1['L2'] = 'Replicate1'
        ws1['M2'] = 'Replicate2'
        ws1['N2'] = 'Replicate3'
        ws1.merge_cells('O1:Q1')
        ws1['O1'] = f'Chain{target_chain} SD4'
        ws1['O2'] = 'Replicate1'
        ws1['P2'] = 'Replicate2'


        count_A = sum(1 for row in ws1['A'] if row.value is not None)
        for row in range(3, 3 + count_A):
            formula = f'=A{row}/1000'  # Change the formula as needed
            ws1[f'B{row}'].value = formula

        ws2 = wb.copy_worksheet(ws1)
        ws2.title = 'WT_RMSD'

        copy_rmsd_data(rmsd_path1, number_of_timepoints, F_or_G, ws1, ws2, mutant)
        copy_rmsd_data(rmsd_path2, number_of_timepoints, F_or_G, ws1, ws2, mutant)
        copy_rmsd_data(rmsd_path3, number_of_timepoints, F_or_G, ws1, ws2, mutant)
        filename = f'{output_path}{mutant}_{F_or_G}_COMPILED.xlsx'
        wb.save(filename)
    if F_or_G == 'Gactin':
        ws1 = wb.active
        ws1.title = f'{mutant}_RMSD'
        rmsd_path1 = f'{output_path}Replicate1/RMSD/'
        rmsd_path2 = f'{output_path}Replicate2/RMSD/'
        rmsd_path3 = f'{output_path}Replicate3/RMSD/'
        copy_time_data_rmsd(rmsd_path1, F_or_G, mutant, ws1, number_of_timepoints, target_chain)
        # add title information for RMSD
        ws1['B2'] = 'Time'
        ws1.merge_cells('C1:E1')
        ws1['C1'] = 'Whole'
        ws1['C2'] = 'Replicate1'
        ws1['D2'] = 'Replicate2'
        ws1['E2'] = 'Replicate3'
        ws1.merge_cells('F1:H1')
        ws1['F1'] = 'SD1'
        ws1['F2'] = 'Replicate1'
        ws1['G2'] = 'Replicate2'
        ws1['H2'] = 'Replicate3'
        ws1.merge_cells('I1:K1')
        ws1['I1'] = 'SD2'
        ws1['I2'] = 'Replicate1'
        ws1['J2'] = 'Replicate2'
        ws1['K2'] = 'Replicate3'
        ws1.merge_cells('L1:N1')
        ws1['L1'] = 'SD3'
        ws1['L2'] = 'Replicate1'
        ws1['M2'] = 'Replicate2'
        ws1['N2'] = 'Replicate3'
        ws1.merge_cells('O1:Q1')
        ws1['O1'] = 'SD4'
        ws1['O2'] = 'Replicate1'
        ws1['P2'] = 'Replicate2'
        ws1['Q2'] = 'Replicate3'

        count_A = sum(1 for row in ws1['A'] if row.value is not None)
        for row in range(3, 3 + count_A):
            formula = f'=A{row}/1000'  # Change the formula as needed
            ws1[f'B{row}'].value = formula

        ws2 = wb.copy_worksheet(ws1)
        ws2.title = 'WT_RMSD'

        copy_rmsd_data(rmsd_path1, number_of_timepoints, F_or_G, ws1, ws2, mutant)
        copy_rmsd_data(rmsd_path2, number_of_timepoints, F_or_G, ws1, ws2, mutant)
        copy_rmsd_data(rmsd_path3, number_of_timepoints, F_or_G, ws1, ws2, mutant)
        filename = f'{output_path}{mutant}_{F_or_G}_COMPILED.xlsx'
        wb.save(filename)


# copies the time data from file number 1 of the processed .txt files into the excel file
def copy_time_data_rmsd(rmsd_path, F_or_G, mutant, destination_ws, number_of_timepoints, target_chain):
    input_text_file = f'{rmsd_path}1_{mutant}_Chain_{target_chain}_protein.txt'
    output_excel_file = f'{rmsd_path}1_{mutant}_Chain_{target_chain}_protein.xlsx'
    # Read the text file into a pandas DataFrame
    data = pd.read_csv(input_text_file, sep='\t')  # Modify 'sep' based on your text file's delimiter
    # Write the DataFrame to an Excel file
    data.to_excel(output_excel_file, index=False)
    rmsd_file_with_time_data = load_workbook(output_excel_file)
    source_ws = rmsd_file_with_time_data.active
    source_range = f'A1:A{number_of_timepoints + 1}'
    destination_cell = 'A3'
    for row_idx, row in enumerate(source_ws[source_range], start=1):
        for col_idx, cell in enumerate(row, start=1):
            dest_cell = destination_ws.cell(row=row_idx + int(destination_cell[1:]) - 1,
                                            column=col_idx + coordinate_to_tuple(destination_cell)[1] - 1)
            dest_cell.value = cell.value


# this file opens all of the processed_...txt files from before, and copies the data to the excel file
def copy_rmsd_data(rmsd_path, number_of_timepoints, F_or_G, mutant_ws, wt_ws, mutant):
    max_lines = number_of_timepoints + 1
    delimiter = '\t'
    # store file numbers for excel calculations
    if F_or_G == 'Gactin':
        mutant_file_offset = 1
        WT_file_offset = 6
    if F_or_G == 'Factin':
        mutant_file_offset = 1
        WT_file_offset = 6
    for filename in os.listdir(rmsd_path):
        if filename.startswith("processed_") and filename.endswith('.txt'):
            if "Replicate1" in rmsd_path:
                start_col = 3
                if mutant in filename:
                    file_number = int(filename.split('_')[1])
                    file_path = os.path.join(rmsd_path, filename)
                    target_column = start_col + (file_number - mutant_file_offset) * 3
                    with open(file_path, 'r') as txt_file:
                        lines = txt_file.readlines()
                        column_C_data = [line.split(delimiter)[2] for line in lines[:max_lines]]
                    for idx, data in enumerate(column_C_data, start=3):
                        mutant_ws.cell(row=idx, column=target_column).value = float(data) if data else None
                if "WT" in filename:
                    file_number = int(filename.split('_')[1])
                    file_path = os.path.join(rmsd_path, filename)
                    target_column = start_col + (file_number - WT_file_offset) * 3
                    with open(file_path, 'r') as txt_file:
                        lines = txt_file.readlines()
                        column_C_data = [line.split(delimiter)[2] for line in lines[:max_lines]]
                    for idx, data in enumerate(column_C_data, start=3):
                        wt_ws.cell(row=idx, column=target_column).value = float(data) if data else None
            if "Replicate2" in rmsd_path:
                start_col = 4
                if mutant in filename:
                    file_number = int(filename.split('_')[1])
                    file_path = os.path.join(rmsd_path, filename)
                    target_column = start_col + (file_number - mutant_file_offset) * 3
                    with open(file_path, 'r') as txt_file:
                        lines = txt_file.readlines()
                        column_C_data = [line.split(delimiter)[2] for line in lines[:max_lines]]
                    for idx, data in enumerate(column_C_data, start=3):
                        mutant_ws.cell(row=idx, column=target_column).value = float(data) if data else None
                if 'WT' in filename:
                    file_number = int(filename.split('_')[1])
                    file_path = os.path.join(rmsd_path, filename)
                    target_column = start_col + (file_number - WT_file_offset) * 3
                    with open(file_path, 'r') as txt_file:
                        lines = txt_file.readlines()
                        column_C_data = [line.split(delimiter)[2] for line in lines[:max_lines]]
                    for idx, data in enumerate(column_C_data, start=3):
                        wt_ws.cell(row=idx, column=target_column).value = float(data) if data else None
            if "Replicate3" in rmsd_path:
                start_col = 5
                if mutant in filename:
                    file_number = int(filename.split('_')[1])
                    file_path = os.path.join(rmsd_path, filename)
                    target_column = start_col + (file_number - mutant_file_offset) * 3
                    with open(file_path, 'r') as txt_file:
                        lines = txt_file.readlines()
                        column_C_data = [line.split(delimiter)[2] for line in lines[:max_lines]]
                    for idx, data in enumerate(column_C_data, start=3):
                        mutant_ws.cell(row=idx, column=target_column).value = float(data) if data else None
                if 'WT' in filename:
                    file_number = int(filename.split('_')[1])
                    file_path = os.path.join(rmsd_path, filename)
                    target_column = start_col + (file_number - WT_file_offset) * 3
                    with open(file_path, 'r') as txt_file:
                        lines = txt_file.readlines()
                        column_C_data = [line.split(delimiter)[2] for line in lines[:max_lines]]
                    for idx, data in enumerate(column_C_data, start=3):
                        wt_ws.cell(row=idx, column=target_column).value = float(data) if data else None

def process_rmsf_files(rmsf_path, F_or_G, mutant, target_chain):
    delimiter = ' '
    file_counter = 1

    naming_phrases = [
        'C-alpha.',
        'C-alpha_EQ.',
        'Protein.',
        'Protein_EQ.'
    ]

    for naming_phrase in naming_phrases:
        for filename in os.listdir(rmsf_path):
            if f'{mutant}_Chain_{target_chain}_rmsf' in filename and '.xvg' in filename and naming_phrase in filename:
                input_file = os.path.join(rmsf_path, filename)
                output_file = os.path.join(rmsf_path, f"processed_{file_counter}_{mutant}_{F_or_G}_{naming_phrase}txt")

                with open(input_file, 'r') as f:
                    lines = f.readlines()
                # Remove lines starting with '#' or '@' (common in .xvg files)
                lines = [line.strip() for line in lines if
                         not line.startswith('#') and not line.startswith('@') and line.strip()]
                lines = [line.strip() for line in lines if
                         not line.startswith('@') and not line.startswith('@') and line.strip()]
                lines = [line.strip() for line in lines if
                         not line.startswith('&') and not line.startswith('&') and line.strip()]
                # Split the lines based on the delimiter and create a list of tuples with two columns
                two_columns = []
                for line in lines:
                    split_line = line.split(delimiter)
                    two_columns.append((split_line[0], ' '.join(split_line[1:])))
                # Calculate the formula (B * 10) and write it to the corresponding cell in column C
                with open(output_file, 'w') as f_out:
                    for idx, (col_A, col_B) in enumerate(two_columns):
                        try:
                            # Convert col_B to float and multiply by 10 if it's not zero
                            value_b = float(col_B)
                            if value_b != 0:
                                value_c = value_b * 10
                            else:
                                value_c = 0  # Keep value as 0 if col_B is zero
                        except ValueError:
                            value_c = 'Invalid'  # Or handle the invalid value case differently
                        # Write the values to the output file (or you can write to Excel using a library like openpyxl)
                        f_out.write(f"{col_A}\t{col_B}\t{value_c}\n")

                file_counter += 1

    for naming_phrase in naming_phrases:
        for filename in os.listdir(rmsf_path):
            if f'WT_Chain_{target_chain}_rmsf' in filename and '.xvg' in filename and naming_phrase in filename:
                input_file = os.path.join(rmsf_path, filename)
                output_file = os.path.join(rmsf_path, f"processed_{file_counter}_WT_{F_or_G}_{naming_phrase}txt")

                with open(input_file, 'r') as f:
                    lines = f.readlines()
                # Remove lines starting with '#' or '@' (common in .xvg files)
                lines = [line.strip() for line in lines if
                         not line.startswith('#') and not line.startswith('@') and line.strip()]
                lines = [line.strip() for line in lines if
                         not line.startswith('@') and not line.startswith('@') and line.strip()]
                lines = [line.strip() for line in lines if
                         not line.startswith('&') and not line.startswith('&') and line.strip()]
                # Split the lines based on the delimiter and create a list of tuples with two columns
                two_columns = []
                for line in lines:
                    split_line = line.split(delimiter)
                    two_columns.append((split_line[0], ' '.join(split_line[1:])))
                # Calculate the formula (B * 10) and write it to the corresponding cell in column C
                with open(output_file, 'w') as f_out:
                    for idx, (col_A, col_B) in enumerate(two_columns):
                        try:
                            # Convert col_B to float and multiply by 10 if it's not zero
                            value_b = float(col_B)
                            if value_b != 0:
                                value_c = value_b * 10
                            else:
                                value_c = 0  # Keep value as 0 if col_B is zero
                        except ValueError:
                            value_c = 'Invalid'  # Or handle the invalid value case differently
                        # Write the values to the output file (or you can write to Excel using a library like openpyxl)
                        f_out.write(f"{col_A}\t{col_B}\t{value_c}\n")

                file_counter += 1


def compile_rmsf(rmsf_path1, rmsf_path2, rmsf_path3, F_or_G, mutant, number_of_residues, path_for_compiled_data, target_chain_number):
    # Load the workbook
    wb = load_workbook(f'{path_for_compiled_data}{mutant}_{F_or_G}_COMPILED.xlsx')
    # Create a new worksheet named 'RMSF'
    ws3 = wb.create_sheet(title='RMSF')
    ws3.merge_cells('C1:N1')
    ws3['C1'] = f'{mutant}'
    ws3.merge_cells('O1:Z1')
    ws3['O1'] = 'WT'

    ws3.merge_cells('C2:E2')
    ws3['C2'] = 'C-alpha'
    ws3.merge_cells('F2:H2')
    ws3['F2'] = 'C-alpha EQ'
    ws3.merge_cells('I2:K2')
    ws3['I2'] = 'Protein'
    ws3.merge_cells('L2:N2')
    ws3['L2'] = 'Protein EQ'
    ws3['B3'] = 'Residue Number'
    ws3['C3'] = 'Replicate1'
    ws3['D3'] = 'Replicate2'
    ws3['E3'] = 'Replicate3'
    ws3['F3'] = 'Replicate1'
    ws3['G3'] = 'Replicate2'
    ws3['H3'] = 'Replicate3'
    ws3['I3'] = 'Replicate1'
    ws3['J3'] = 'Replicate2'
    ws3['K3'] = 'Replicate3'
    ws3['L3'] = 'Replicate1'
    ws3['M3'] = 'Replicate2'
    ws3['N3'] = 'Replicate3'

    ws3.merge_cells('O2:Q2')
    ws3['O2'] = 'C-alpha'
    ws3.merge_cells('R2:T2')
    ws3['R2'] = 'C-alpha EQ'
    ws3.merge_cells('U2:W2')
    ws3['U2'] = 'Protein'
    ws3.merge_cells('X2:Z2')
    ws3['X2'] = 'Protein EQ'
    ws3['O3'] = 'Replicate1'
    ws3['P3'] = 'Replicate2'
    ws3['Q3'] = 'Replicate3'
    ws3['R3'] = 'Replicate1'
    ws3['S3'] = 'Replicate2'
    ws3['T3'] = 'Replicate3'
    ws3['U3'] = 'Replicate1'
    ws3['V3'] = 'Replicate2'
    ws3['W3'] = 'Replicate3'
    ws3['X3'] = 'Replicate1'
    ws3['Y3'] = 'Replicate2'
    ws3['Z3'] = 'Replicate3'


    # Copy headers to other columns for individual chains
    source_range = ws3['B1:Z3']
    copied_data = []
    for row in source_range:
        copied_data.append([cell.value for cell in row])

    # Paste the copied values to the destination range (AB1)
    for row_idx, row in enumerate(copied_data, start=1):
        for col_idx, value in enumerate(row, start=1):
            col_value = 28
            ws3.cell(row=row_idx, column=col_idx + col_value).value = value

    input_text_file = f'{rmsf_path1}processed_1_{mutant}_{F_or_G}_C-alpha.txt'
    output_excel_file = f'{rmsf_path1}processed_1_{mutant}_{F_or_G}_C-alpha.xlsx'
    # Read the text file into a pandas DataFrame
    data = pd.read_csv(input_text_file, sep='\t')  # Modify 'sep' based on your text file's delimiter
    # Write the DataFrame to an Excel file
    data.to_excel(output_excel_file, index=False)
    source_wb = load_workbook(output_excel_file)
    source_ws = source_wb.active
    # copy residue number from data file
    for row_idx, row in enumerate(source_ws[f'A1:A{number_of_residues + 1}'], start=1):
        for col_idx, cell in enumerate(row, start=1):
            dest_cell = ws3.cell(row=row_idx + int('B4'[1:]) - 1,
                                            column=coordinate_to_tuple('B4')[1])
            dest_cell.value = cell.value

    # for each processed txt file, isolate file number and copy column C to corresponding column
    for filename in os.listdir(rmsf_path1):
        if mutant in filename:
            start_col = 3
            file_offset = 1
        if 'WT' in filename:
            start_col = 15
            file_offset = 5
        if filename.startswith("processed_") and filename.endswith('.txt'):
            file_path = os.path.join(rmsf_path1, filename)
            file_number = int(filename.split('_')[1])
            delimiter = '\t'
            max_lines = number_of_residues + 1
            with open(file_path, 'r') as txt_file:
                lines = txt_file.readlines()
                column_C_data = [line.split(delimiter)[2] for line in
                                 lines[:max_lines]]  # Assuming data is space-separated
            target_column = start_col + (file_number - file_offset) * 3
            for idx, data in enumerate(column_C_data, start=4):
                ws3.cell(row=idx, column=target_column).value = float(data) if data else None

    for filename in os.listdir(rmsf_path2):
        if mutant in filename:
            start_col = 4
            file_offset = 1
        if 'WT' in filename:
            start_col = 16
            file_offset = 5
        if filename.startswith("processed_") and filename.endswith('.txt'):
            file_path = os.path.join(rmsf_path2, filename)
            file_number = int(filename.split('_')[1])
            delimiter = '\t'
            max_lines = number_of_residues + 1
            with open(file_path, 'r') as txt_file:
                lines = txt_file.readlines()
                column_C_data = [line.split(delimiter)[2] for line in
                                 lines[:max_lines]]  # Assuming data is space-separated
            target_column = start_col + (file_number - file_offset) * 3
            for idx, data in enumerate(column_C_data, start=4):
                ws3.cell(row=idx, column=target_column).value = float(data) if data else None

    for filename in os.listdir(rmsf_path3):
        if mutant in filename:
            start_col = 5
            file_offset = 1
        if 'WT' in filename:
            start_col = 17
            file_offset = 5
        if filename.startswith("processed_") and filename.endswith('.txt'):
            file_path = os.path.join(rmsf_path3, filename)
            file_number = int(filename.split('_')[1])
            delimiter = '\t'
            max_lines = number_of_residues + 1
            with open(file_path, 'r') as txt_file:
                lines = txt_file.readlines()
                column_C_data = [line.split(delimiter)[2] for line in
                                 lines[:max_lines]]  # Assuming data is space-separated
            target_column = start_col + (file_number - file_offset) * 3
            for idx, data in enumerate(column_C_data, start=4):
                ws3.cell(row=idx, column=target_column).value = float(data) if data else None

    #if Factin, isolate chain C data
    if F_or_G == 'Factin':
        target_chain_start_row = ( target_chain_number * 375 ) + 4
        target_chain_end_row = target_chain_start_row + 374
        cell_range = {'AC4': f'B{target_chain_start_row}:Z{target_chain_end_row}'}
        for start_cell, source_range in cell_range.items():
            source_rows = ws3[source_range]
            target_cell = ws3[start_cell]

            for dest_row, source_row in enumerate(source_rows, start=1):
                for col_idx, cell in enumerate(source_row, start=1):
                    dest_cell = ws3.cell(row=target_cell.row + dest_row - 1, column=target_cell.column + col_idx - 1)
                    dest_cell.value = cell.value
    wb.save(f'{path_for_compiled_data}{mutant}_{F_or_G}_COMPILED.xlsx')


def process_PCA(PCA_path, F_or_G, mutant):
    delimiter = ' '

    for filename in os.listdir(PCA_path):
        if 'eigenrmsf' in filename:
            if mutant in filename:
                input_file = os.path.join(PCA_path, filename)
                output_file = os.path.join(PCA_path, f"processed_{mutant}_{F_or_G}_PCA_RMSF.txt")

                with open(input_file, 'r') as f:
                    lines = f.readlines()
                # Remove lines starting with '#' or '@' (common in .xvg files)
                lines = [line.strip() for line in lines if
                         not line.startswith('#') and not line.startswith('@') and line.strip()]
                lines = [line.strip() for line in lines if
                         not line.startswith('@') and not line.startswith('@') and line.strip()]
                lines = [line.strip() for line in lines if
                         not line.startswith('&') and not line.startswith('&') and line.strip()]
                # Split the lines based on the delimiter and create a list of tuples with two columns
                two_columns = []
                for line in lines:
                    split_line = line.split(delimiter)
                    two_columns.append((split_line[0], ' '.join(split_line[1:])))
                # Calculate the formula (B * 10) and write it to the corresponding cell in column C
                with open(output_file, 'w') as f_out:
                    for idx, (col_A, col_B) in enumerate(two_columns):
                        try:
                            # Convert col_B to float and multiply by 10 if it's not zero
                            value_b = float(col_B)
                            if value_b != 0:
                                value_c = value_b * 10
                            else:
                                value_c = 0  # Keep value as 0 if col_B is zero
                        except ValueError:
                            value_c = 'Invalid'  # Or handle the invalid value case differently
                        # Write the values to the output file (or you can write to Excel using a library like openpyxl)
                        f_out.write(f"{col_A}\t{col_B}\t{value_c}\n")
            if 'WT' in filename:
                input_file = os.path.join(PCA_path, filename)
                output_file = os.path.join(PCA_path, f"processed_WT_{F_or_G}_PCA_RMSF.txt")

                with open(input_file, 'r') as f:
                    lines = f.readlines()
                # Remove lines starting with '#' or '@' (common in .xvg files)
                lines = [line.strip() for line in lines if
                         not line.startswith('#') and not line.startswith('@') and line.strip()]
                lines = [line.strip() for line in lines if
                         not line.startswith('@') and not line.startswith('@') and line.strip()]
                lines = [line.strip() for line in lines if
                         not line.startswith('&') and not line.startswith('&') and line.strip()]
                # Split the lines based on the delimiter and create a list of tuples with two columns
                two_columns = []
                for line in lines:
                    split_line = line.split(delimiter)
                    two_columns.append((split_line[0], ' '.join(split_line[1:])))
                # Calculate the formula (B * 10) and write it to the corresponding cell in column C
                with open(output_file, 'w') as f_out:
                    for idx, (col_A, col_B) in enumerate(two_columns):
                        try:
                            # Convert col_B to float and multiply by 10 if it's not zero
                            value_b = float(col_B)
                            if value_b != 0:
                                value_c = value_b * 10
                            else:
                                value_c = 0  # Keep value as 0 if col_B is zero
                        except ValueError:
                            value_c = 'Invalid'  # Or handle the invalid value case differently
                        # Write the values to the output file (or you can write to Excel using a library like openpyxl)
                        f_out.write(f"{col_A}\t{col_B}\t{value_c}\n")

    for filename in os.listdir(PCA_path):
        if 'processed_' in filename:
            input_file_path = os.path.join(PCA_path, filename)
            # Read the specified .txt file into a DataFrame
            df = pd.read_csv(input_file_path, header=None, names=['Column A'])
            # Count the number of non-empty lines in column A
            number_of_lines = len(df[df['Column A'].notnull()])

            start_line1 = (number_of_lines // 2) + 1
            start_line2 = 1
            end_line1 = number_of_lines
            end_line2 = number_of_lines // 2
            name_parts = filename.split('_')  # Split the filename by '_'
            if len(name_parts) >= 2:
                output_name_part = name_parts[1].split('.')[0]  # Get the part after the first '_'
                # Generate the output file name using the extracted part
                output_file_path_PC1 = os.path.join(PCA_path, f"PC1_{output_name_part}.txt")
                with open(input_file_path, 'r') as file:
                    lines = file.readlines()
                updated_lines_PC1 = [line for i, line in enumerate(lines) if i + 1 < start_line1 or i + 1 > end_line1]
                with open(output_file_path_PC1, 'w') as file:
                    file.writelines(updated_lines_PC1)

                output_file_path_PC2 = os.path.join(PCA_path, f"PC2_{output_name_part}.txt")
                with open(input_file_path, 'r') as file:
                    lines = file.readlines()
                # Filter out lines within the specified range to be deleted
                updated_lines_PC2 = [line for i, line in enumerate(lines) if i + 1 < start_line2 or i + 1 > end_line2]
                with open(output_file_path_PC2, 'w') as file:
                    file.writelines(updated_lines_PC2)


def compile_PCA(path_for_compiled_data, F_or_G, mutant, PCA_path, number_of_residues):
    # Load the workbook
    wb = load_workbook(f'{path_for_compiled_data}{mutant}_{F_or_G}_COMPILED.xlsx')
    # Create a new worksheet
    ws4 = wb.create_sheet(title='PC1_rmsf')
    ws4['B3'] = 'Residue Number'
    
    ws4['C2'] = f'{mutant}'
    ws4['C3'] = 'Combined Replicates'
    ws4['D2'] = 'WT'
    ws4['D3'] = 'Combined Replicates'


    # make columns for deltaRMSF
################################################NEED TO CHANGE IF DIFFERENT FILAMENT LENGTH THAN 5##################
    ws4['J3'] = 'DeltaRMSF'
    ws4['L3'] = 'Residue Number'
    ws4['M3'] = 'ChainA'
    ws4['N3'] = 'ChainB'
    ws4['O3'] = 'ChainC'
    ws4['P3'] = 'ChainD'
    ws4['Q3'] = 'ChainE'

    # make columns for all POSITIVE deltaRMSF (needed for plot formatting)
    ws4['S3'] = 'Residue Number'
    ws4['T3'] = 'ChainA'
    ws4['U3'] = 'ChainB'
    ws4['V3'] = 'ChainC'
    ws4['W3'] = 'ChainD'
    ws4['X3'] = 'ChainE'

    # make columns for all NEGATIVE deltaRMSF (needed for plot formatting)
    ws4['Z3'] = 'Residue Number'
    ws4['AA3'] = 'ChainA'
    ws4['AB3'] = 'ChainB'
    ws4['AC3'] = 'ChainC'
    ws4['AD3'] = 'ChainD'
    ws4['AE3'] = 'ChainE'


    ws3 = wb['RMSF']
    # copy over residue numbers from RMSF sheet (column B or column 2 in ws3)
    for row_index in range(1, ws3.max_row + 1):
        # Get the value from the source column and assign it to the destination column
        ws4.cell(row=row_index, column=2).value = ws3.cell(row=row_index, column=2).value

    # insert formulas for deltaRMSF
    # Calculate the max row in column B (assuming it contains data)
    max_row1 = ws4.max_row
    # Loop through rows starting from row 4 in column B
    for row in range(4, max_row1 + 1):
        # Create the formula and insert it into column H for each row
        formula = f"=D{row} - C{row}"
        ws4[f'J{row}'] = formula

    # set formula to calculate positive deltaRMSF
    for row in range(4, 379):  # Rows 4 to 378
        for col in range(13, 18):  # columns M to Q for the positive RMSF calculations (column indexes 13-17)
            cell_t = ws4.cell(row=row, column=col + 7)  #set column to copy to (column t)
            cell_m = ws4.cell(row=row, column=col) #set origin column
            formula = f'=IF({cell_m.coordinate}>=0, {cell_m.coordinate}, 0)'  # Create the formula
            cell_t.value = formula  # Assign the formula to the cell

    # set formula to calculate negative deltaRMSF
    for row in range(4, 379):  # Rows 4 to 378
        for col in range(13, 18):
            cell_aa = ws4.cell(row=row, column=col + 14) #set column to copy to (column AA
            cell_m = ws4.cell(row=row, column=col)
            formula = f'=IF({cell_m.coordinate}<=0, {cell_m.coordinate}, 0)'  # Create the formula
            cell_aa.value = formula  # Assign the formula to the cell

    ws5 = wb.copy_worksheet(ws4)
    ws5.title = 'PC2_rmsf'

    # copy PCA_rmsf information to excel
    for filename in os.listdir(PCA_path):
        if filename.startswith(f"PC1_{mutant}") and filename.endswith('.txt'):
            destination_sheet = ws4
            delimiter = '\t'
            max_lines = number_of_residues + 1
            file_path = os.path.join(PCA_path, filename)
            with open(file_path, 'r') as txt_file:
                lines = txt_file.readlines()
                column_C_data = [line.split(delimiter)[2] for line in
                                 lines[:max_lines]]  # Assuming data is space-separated
            target_column = 3
            for idx, data in enumerate(column_C_data, start=4):
                destination_sheet.cell(row=idx, column=target_column).value = float(data) if data else None
        if filename.startswith(f"PC1_WT") and filename.endswith('.txt'):
            destination_sheet = ws4
            delimiter = '\t'
            max_lines = number_of_residues + 1
            file_path = os.path.join(PCA_path, filename)
            with open(file_path, 'r') as txt_file:
                lines = txt_file.readlines()
                column_C_data = [line.split(delimiter)[2] for line in
                                 lines[:max_lines]]  # Assuming data is space-separated
            target_column = 4
            for idx, data in enumerate(column_C_data, start=4):
                destination_sheet.cell(row=idx, column=target_column).value = float(data) if data else None
        if filename.startswith(f"PC2_{mutant}") and filename.endswith('.txt'):
            destination_sheet = ws5
            delimiter = '\t'
            max_lines = number_of_residues + 1
            file_path = os.path.join(PCA_path, filename)
            with open(file_path, 'r') as txt_file:
                lines = txt_file.readlines()
                column_C_data = [line.split(delimiter)[2] for line in
                                 lines[:max_lines]]  # Assuming data is space-separated
            target_column = 3
            for idx, data in enumerate(column_C_data, start=4):
                destination_sheet.cell(row=idx, column=target_column).value = float(data) if data else None
        if filename.startswith(f"PC2_WT") and filename.endswith('.txt'):
            destination_sheet = ws5
            delimiter = '\t'
            max_lines = number_of_residues + 1
            file_path = os.path.join(PCA_path, filename)
            with open(file_path, 'r') as txt_file:
                lines = txt_file.readlines()
                column_C_data = [line.split(delimiter)[2] for line in
                                 lines[:max_lines]]  # Assuming data is space-separated
            target_column = 4
            for idx, data in enumerate(column_C_data, start=4):
                destination_sheet.cell(row=idx, column=target_column).value = float(data) if data else None

    # Once PCA_rmsf information is pasted, separate by individual chains
    # Define cell locations and corresponding ranges
    cell_ranges = {
        'L4': 'B4:B378',
        'S4': 'B4:B378',
        'Z4': 'B4:B378',
        'M4': 'J4:J378',
        'N4': 'J379:J753',
        'O4': 'J754:J1128',
        'P4': 'J1129:J1503',
        'Q4': 'J1504:J1878'
    }
    # Iterate over cell_ranges and copy data to specified cells
    for start_cell, source_range in cell_ranges.items():
        source_rows = ws4[source_range]
        target_cell = ws4[start_cell]

        for dest_row, source_row in enumerate(source_rows, start=1):
            for col_idx, cell in enumerate(source_row, start=1):
                dest_cell = ws4.cell(row=target_cell.row + dest_row - 1, column=target_cell.column + col_idx - 1)
                dest_cell.value = cell.value
    # Iterate over cell_ranges and copy data to specified cells
    for start_cell, source_range in cell_ranges.items():
        source_rows = ws5[source_range]
        target_cell = ws5[start_cell]

        for dest_row, source_row in enumerate(source_rows, start=1):
            for col_idx, cell in enumerate(source_row, start=1):
                dest_cell = ws5.cell(row=target_cell.row + dest_row - 1, column=target_cell.column + col_idx - 1)
                dest_cell.value = cell.value

    wb.save(f'{path_for_compiled_data}{mutant}_{F_or_G}_COMPILED.xlsx')


def process_HBOND_files(folder_path, F_or_G, mutant, chains, target_chain_number, nucleotide, target_chain):
    if F_or_G == 'Factin':
        protein_index_offset = 6
        nucleotide_index_offset = 5

        protein_array = [
            f"hbonds_Chain_{target_chain}_Chain_{target_chain}",
            f"hbonds_Chain_{target_chain}_SD1_Chain_{target_chain}_SD1",
            f"hbonds_Chain_{target_chain}_SD2_Chain_{target_chain}_SD2",
            f"hbonds_Chain_{target_chain}_SD3_Chain_{target_chain}_SD3",
            f"hbonds_Chain_{target_chain}_SD4_Chain_{target_chain}_SD4"
        ]

        nucleotide_array = [
            f"hbonds_{nucleotide}_Chain_{target_chain}_SD1",
            f"hbonds_{nucleotide}_Chain_{target_chain}_SD2",
            f"hbonds_{nucleotide}_Chain_{target_chain}_SD3",
            f"hbonds_{nucleotide}_Chain_{target_chain}_SD4"
        ]

    if F_or_G == 'Gactin':
        protein_index_offset = 6
        nucleotide_index_offset = 5

        protein_array = [
            f"hbonds_Chain_{target_chain}_Chain_{target_chain}",
            f"hbonds_Chain_{target_chain}_SD1_Chain_{target_chain}_SD1",
            f"hbonds_Chain_{target_chain}_SD2_Chain_{target_chain}_SD2",
            f"hbonds_Chain_{target_chain}_SD3_Chain_{target_chain}_SD3",
            f"hbonds_Chain_{target_chain}_SD4_Chain_{target_chain}_SD4"
        ]

        nucleotide_array = [
            f"hbonds_{nucleotide}_Chain_{target_chain}_SD1",
            f"hbonds_{nucleotide}_Chain_{target_chain}_SD2",
            f"hbonds_{nucleotide}_Chain_{target_chain}_SD3",
            f"hbonds_{nucleotide}_Chain_{target_chain}_SD4"
        ]

    for idx, text_string in enumerate(protein_array):
        file_list = [file for file in os.listdir(folder_path) if file.endswith('.xvg') and text_string in file]

        for filename in file_list:
            file_path = os.path.join(folder_path, filename)

            # Read the file and process lines according to the specified conditions
            with open(file_path, 'r') as file:
                lines = file.readlines()
                lines = [line.strip() for line in lines if
                         not line.startswith('#') and not line.startswith('@') and line.strip()]
                lines = [line.strip() for line in lines if
                         not line.startswith('@') and not line.startswith('@') and line.strip()]
                lines = [line.strip() for line in lines if
                         not line.startswith('&') and not line.startswith('&') and line.strip()]

            # Replace consecutive spaces with a single space before splitting
            lines = [' '.join(line.split()) for line in lines]

            # Split lines using space as delimiter
            lines = [line.split(' ') for line in lines]

            if f'{mutant}' in filename:
                # Save the processed data to a new .txt file
                output_filename = f"processed_{idx + 1}_{filename.split('.')[0]}.txt"
                output_file_path = os.path.join(folder_path, output_filename)
                with open(output_file_path, 'w') as output_file:
                    for line in lines:
                        output_file.write('\t'.join(line) + '\n')  # Use tab as delimiter for .txt file

            if "WT" in filename:
                # Save the processed data to a new .txt file
                output_filename = f"processed_{idx + protein_index_offset}_{filename.split('.')[0]}.txt"
                output_file_path = os.path.join(folder_path, output_filename)
                with open(output_file_path, 'w') as output_file:
                    for line in lines:
                        output_file.write('\t'.join(line) + '\n')  # Use tab as delimiter for .txt file

    for idx, text_string in enumerate(nucleotide_array):
        file_list = [file for file in os.listdir(folder_path) if file.endswith('.xvg') and text_string in file]

        for filename in file_list:
            file_path = os.path.join(folder_path, filename)

            # Read the file and process lines according to the specified conditions
            with open(file_path, 'r') as file:
                lines = file.readlines()
                lines = [line.strip() for line in lines if
                         not line.startswith('#') and not line.startswith('@') and line.strip()]
                lines = [line.strip() for line in lines if
                         not line.startswith('@') and not line.startswith('@') and line.strip()]
                lines = [line.strip() for line in lines if
                         not line.startswith('&') and not line.startswith('&') and line.strip()]

            # Replace consecutive spaces with a single space before splitting
            lines = [' '.join(line.split()) for line in lines]

            # Split lines using space as delimiter
            lines = [line.split(' ') for line in lines]

            if f'{mutant}' in filename:
                # Save the processed data to a new .txt file
                output_filename = f"processed_{idx + 1}_{filename.split('.')[0]}.txt"
                output_file_path = os.path.join(folder_path, output_filename)
                with open(output_file_path, 'w') as output_file:
                    for line in lines:
                        output_file.write('\t'.join(line) + '\n')  # Use tab as delimiter for .txt file

            if "WT" in filename:
                # Save the processed data to a new .txt file
                output_filename = f"processed_{idx + nucleotide_index_offset}_{filename.split('.')[0]}.txt"
                output_file_path = os.path.join(folder_path, output_filename)
                with open(output_file_path, 'w') as output_file:
                    for line in lines:
                        output_file.write('\t'.join(line) + '\n')  # Use tab as delimiter for .txt file


def compile_hbonds(F_or_G, mutant, path_for_compiled_data, hbonds_rep1_path, hbonds_rep2_path, hbonds_rep3_path,
                   chains, target_chain_number, nucleotide):

    # first prepare excel sheets
    target_chain = chains[target_chain_number]
    if F_or_G == 'Factin':
        max_column = 33
        protein_index_offset = 6
        nucleotide_index_offset = 5
        # Create an array of text strings
        header_labels = [
            f"Chain {target_chain}",
            f"Chain {target_chain} SD1",
            f"Chain {target_chain} SD2",
            f"Chain {target_chain} SD3",
            f"Chain {target_chain} SD4"
        ]

    if F_or_G == 'Gactin':
        max_column = 33
        protein_index_offset = 6
        nucleotide_index_offset = 5
        # Create an array of text strings
        header_labels = [
            "Whole",
            "SD1",
            "SD2",
            "SD3",
            "SD4"
        ]

    wb = load_workbook(f'{path_for_compiled_data}{mutant}_{F_or_G}_COMPILED.xlsx')
    # Create a new worksheet
    ws6 = wb.create_sheet(title='hbonds_protein')
    ws1 = wb[f'{mutant}_RMSD']
    # Assign text_strings values to individual cells before merging
    col = 3
    for header_label in header_labels:
        cell = ws6.cell(row=1, column=col)
        cell.value = header_label
        col += 6

    # Merge cells after assigning values
    for col in range(3, max_column, 6):
        merge_range = f'{ws6.cell(row=1, column=col).coordinate}:{ws6.cell(row=1, column=col + 5).coordinate}'
        ws6.merge_cells(merge_range)

    # Assign alternating values "3er" and "WT" to individual cells before merging
    for col in range(3, max_column, 6):
        cell = ws6.cell(row=2, column=col)
        cell.value = f'{mutant}'
        col += 6

    for col in range(6, max_column, 6):
        cell = ws6.cell(row=2, column=col)
        cell.value = 'WT'
        col += 6

    # Merge cells after assigning values
    for col in range(3, max_column, 3):
        merge_range = f'{ws6.cell(row=2, column=col).coordinate}:{ws6.cell(row=2, column=col + 2).coordinate}'
        ws6.merge_cells(merge_range)

    # Assign "Replicate2" to even cells (if col value divided by 2 is true aka 0), otherwise "Replicate1"
    for col in range(3, max_column + 2, 3):
        cell = ws6.cell(row=3, column=col)
        cell.value = "Replicate1"
        col += 1
        cell = ws6.cell(row=3, column=col)
        cell.value = "Replicate2"
        col += 1
        cell = ws6.cell(row=3, column=col)
        cell.value = "Replicate3"
        col += 1

    # copy time from RMSD sheet
    for row in range(2, 10004):
        source_cell = ws1.cell(row=row, column=1)
        value = source_cell.value  # Get the calculated value from the formula
        # Paste the value into the destination sheet
        destination_cell = ws6.cell(row=row + 1, column=1)
        destination_cell.value = value

    count_A = sum(1 for row in ws1['A'] if row.value is not None)
    for row in range(4, 4 + count_A):
        formula = f'=A{row}/1000'  # Change the formula as needed
        ws6[f'B{row}'].value = formula

    ws7 = wb.copy_worksheet(ws6)
    ws7.title = 'hbonds_nucleotide'

    # Iterate through each .txt file in the directory
    for filename in sorted(os.listdir(hbonds_rep1_path)):
        if filename.startswith('processed_') and filename.endswith('.txt') and f'{nucleotide}' not in filename:
            file_number = int(filename.split('_')[1].split('.')[0])  # Extract the file number

            # Read the contents of column B from the .txt file
            with open(os.path.join(hbonds_rep1_path, filename), 'r') as file:
                column_B_data = [line.split()[1] for line in file.readlines()]

            if mutant in filename:
                # Calculate the column to paste data (every 4th column starting from target_column)
                paste_column = 3 + 6 * (file_number - 1)
                # Paste the data into the specified column
                for i, value in enumerate(column_B_data, start=1):
                    ws6.cell(row=i + 3, column=paste_column, value=value)
            if 'WT' in filename:
                # Calculate the column to paste data (every 4th column starting from target_column)
                paste_column = 6 + 6 * (file_number - protein_index_offset)
                # Paste the data into the specified column
                for i, value in enumerate(column_B_data, start=1):
                    ws6.cell(row=i + 3, column=paste_column, value=value)
    for filename in sorted(os.listdir(hbonds_rep2_path)):
        if filename.startswith('processed_') and filename.endswith('.txt') and f'{nucleotide}' not in filename:
            file_number = int(filename.split('_')[1].split('.')[0])  # Extract the file number

            # Read the contents of column B from the .txt file
            with open(os.path.join(hbonds_rep2_path, filename), 'r') as file:
                column_B_data = [line.split()[1] for line in file.readlines()]

            if mutant in filename:
                # Calculate the column to paste data (every 4th column starting from target_column)
                paste_column = 4 + 6 * (file_number - 1)
                # Paste the data into the specified column
                for i, value in enumerate(column_B_data, start=1):
                    ws6.cell(row=i + 3, column=paste_column, value=value)
            if 'WT' in filename:
                # Calculate the column to paste data (every 4th column starting from target_column)
                paste_column = 7 + 6 * (file_number - protein_index_offset)
                # Paste the data into the specified column
                for i, value in enumerate(column_B_data, start=1):
                    ws6.cell(row=i + 3, column=paste_column, value=value)
    for filename in sorted(os.listdir(hbonds_rep3_path)):
        if filename.startswith('processed_') and filename.endswith('.txt') and f'{nucleotide}' not in filename:
            file_number = int(filename.split('_')[1].split('.')[0])  # Extract the file number

            # Read the contents of column B from the .txt file
            with open(os.path.join(hbonds_rep3_path, filename), 'r') as file:
                column_B_data = [line.split()[1] for line in file.readlines()]

            if mutant in filename:
                # Calculate the column to paste data (every 4th column starting from target_column)
                paste_column = 5 + 6 * (file_number - 1)
                # Paste the data into the specified column
                for i, value in enumerate(column_B_data, start=1):
                    ws6.cell(row=i + 3, column=paste_column, value=value)
            if 'WT' in filename:
                # Calculate the column to paste data (every 4th column starting from target_column)
                paste_column = 8 + 6 * (file_number - protein_index_offset)
                # Paste the data into the specified column
                for i, value in enumerate(column_B_data, start=1):
                    ws6.cell(row=i + 3, column=paste_column, value=value)

    for filename in sorted(os.listdir(hbonds_rep1_path)):
        if filename.startswith('processed_') and filename.endswith('.txt') and f'{nucleotide}' in filename:
            file_number = int(filename.split('_')[1].split('.')[0])  # Extract the file number

            # Read the contents of column B from the .txt file
            with open(os.path.join(hbonds_rep1_path, filename), 'r') as file:
                column_B_data = [line.split()[1] for line in file.readlines()]

            if mutant in filename:
                paste_column = 9 + 6 * (file_number - 1)
                # Paste the data into the specified column
                for i, value in enumerate(column_B_data, start=1):
                    ws7.cell(row=i + 3, column=paste_column, value=value)
            if 'WT' in filename:
                paste_column = 12 + 6 * (file_number - nucleotide_index_offset)
                # Paste the data into the specified column
                for i, value in enumerate(column_B_data, start=1):
                    ws7.cell(row=i + 3, column=paste_column, value=value)
    for filename in sorted(os.listdir(hbonds_rep2_path)):
        if filename.startswith('processed_') and filename.endswith('.txt') and f'{nucleotide}' in filename:
            file_number = int(filename.split('_')[1].split('.')[0])  # Extract the file number

            # Read the contents of column B from the .txt file
            with open(os.path.join(hbonds_rep2_path, filename), 'r') as file:
                column_B_data = [line.split()[1] for line in file.readlines()]

            if mutant in filename:
                paste_column = 10 + 6 * (file_number - 1)
                # Paste the data into the specified column
                for i, value in enumerate(column_B_data, start=1):
                    ws7.cell(row=i + 3, column=paste_column, value=value)
            if 'WT' in filename:
                paste_column = 13 + 6 * (file_number - nucleotide_index_offset)
                # Paste the data into the specified column
                for i, value in enumerate(column_B_data, start=1):
                    ws7.cell(row=i + 3, column=paste_column, value=value)

    for filename in sorted(os.listdir(hbonds_rep3_path)):
        if filename.startswith('processed_') and filename.endswith('.txt') and f'{nucleotide}' in filename:
            file_number = int(filename.split('_')[1].split('.')[0])  # Extract the file number

            # Read the contents of column B from the .txt file
            with open(os.path.join(hbonds_rep3_path, filename), 'r') as file:
                column_B_data = [line.split()[1] for line in file.readlines()]

            if mutant in filename:
                paste_column = 11 + 6 * (file_number - 1)
                # Paste the data into the specified column
                for i, value in enumerate(column_B_data, start=1):
                    ws7.cell(row=i + 3, column=paste_column, value=value)
            if 'WT' in filename:
                paste_column = 14 + 6 * (file_number - nucleotide_index_offset)
                # Paste the data into the specified column
                for i, value in enumerate(column_B_data, start=1):
                    ws7.cell(row=i + 3, column=paste_column, value=value)

    for row in range(4, ws7.max_row + 1):
        if ws7[f'B{row}'].value:
            # Insert formula to add all subdomain Hbonds
            for column in range(3, 9):
                ws7.cell(row=row, column=column).value = f'=SUM({ws7.cell(row=row, column=column + 6).coordinate}, ' \
                                                         f'{ws7.cell(row=row, column=column + 12).coordinate}, ' \
                                                         f'{ws7.cell(row=row, column=column + 18).coordinate}, ' \
                                                         f'{ws7.cell(row=row, column=column + 24).coordinate})'

    wb.save(f'{path_for_compiled_data}{mutant}_{F_or_G}_COMPILED.xlsx')


# process_rmsd_dist_files calls format_rmsd_dist_files for each file in a folder and spits out an output file
# interprotomer rmsd=C-terminus of one chain+D-loop of another
def process_rmsd_dist_files(F_or_G, mutant, interprotomer_folder_path, nucleotide_folder_path, chains, target_chain_number):
    #identify interacting chains with the target
    #note: index numbering in 'chains' starts at 0.
    interprotomer_chain1 = chains[target_chain_number - 2]
    interprotomer_chain2 = chains[target_chain_number + 2]
    target_chain = chains[target_chain_number]
    if F_or_G == 'Factin':
        interprotomer_naming_phrases = [
            f'{mutant}_{F_or_G}_Chain_{interprotomer_chain1}_{target_chain}',
            f'{mutant}_{F_or_G}_Chain_{target_chain}_{interprotomer_chain2}',
            f'WT_{F_or_G}_Chain_{interprotomer_chain1}_{target_chain}',
            f'WT_{F_or_G}_Chain_{target_chain}_{interprotomer_chain2}'
        ]

        nucleotide_naming_phrases = [
            f'{mutant}_{F_or_G}_Chain_{target_chain}_ADP',
            f'{mutant}_{F_or_G}_Chain_{target_chain}_PO4',
            f'WT_{F_or_G}_Chain_{target_chain}_ADP',
            f'WT_{F_or_G}_Chain_{target_chain}_PO4'
        ]

    if F_or_G == 'Gactin':
        nucleotide_naming_phrases = [
            f'{mutant}_{F_or_G}_protein_ADP',
            f'{mutant}_{F_or_G}_protein_PO4',
            f'WT_{F_or_G}_protein_ADP',
            f'WT_{F_or_G}_protein_PO4'
        ]

    file_counter = 1
    delimiter = ' '
    # first, open each .xvg file, strip non-data lines and save it in new format
    if F_or_G == 'Factin':
        for interprotomer_naming_phrase in interprotomer_naming_phrases:
            for filename in os.listdir(interprotomer_folder_path):
                if interprotomer_naming_phrase in filename:
                    input_file = os.path.join(interprotomer_folder_path, filename)
                    output_file = os.path.join(interprotomer_folder_path,
                                               f"{file_counter}_{interprotomer_naming_phrase}_rmsd.txt")
                    with open(input_file, 'r') as f:
                        lines = f.readlines()

                    # Remove lines starting with '#' or '@' (common in .xvg files)
                    lines = [line.strip() for line in lines if
                             not line.startswith('#') and not line.startswith('@') and line.strip()]

                    two_columns = []
                    for line in lines:
                        split_line = line.split(delimiter)
                        two_columns.extend([(split_line[0], ' '.join(split_line[1:]))])

                    with open(output_file, 'w') as f_out:
                        for row in two_columns:
                            f_out.write('\t'.join(row) + '\n')
                    file_counter += 1

        for filename in os.listdir(interprotomer_folder_path):
            if filename.endswith('.txt'):  # Process only .txt files
                input_file = os.path.join(interprotomer_folder_path, filename)
                output_file = os.path.join(interprotomer_folder_path, f"processed_{filename}")
                # Read the contents of the .txt file
                with open(input_file, 'r') as txt_file:
                    lines = txt_file.readlines()

                # Apply formula to multiply values by 10 (if the value is not 0)
                modified_lines = []
                for line in lines:
                    line_data = line.strip().split('\t')  # Assuming tab-separated data, modify delimiter if needed
                    if len(line_data) >= 2:
                        try:
                            value = float(line_data[1])
                            if value != 0:
                                line_data.append(str(value * 10))
                            else:
                                line_data.append('0')
                        except ValueError:
                            line_data.append('Invalid')  # Or handle the invalid value case differently

                    modified_lines.append('\t'.join(line_data))

                # Write the modified data to a new output file
                with open(output_file, 'w') as output_txt_file:
                    output_txt_file.write('\n'.join(modified_lines))

    file_counter = 1
    for nucleotide_naming_phrase in nucleotide_naming_phrases:
        for filename in os.listdir(nucleotide_folder_path):
            if nucleotide_naming_phrase in filename:
                input_file = os.path.join(nucleotide_folder_path, filename)
                output_file = os.path.join(nucleotide_folder_path,
                                           f"{file_counter}_{nucleotide_naming_phrase}_rmsd.txt")
                with open(input_file, 'r') as f:
                    lines = f.readlines()

                # Remove lines starting with '#' or '@' (common in .xvg files)
                lines = [line.strip() for line in lines if
                         not line.startswith('#') and not line.startswith('@') and line.strip()]

                two_columns = []
                for line in lines:
                    split_line = line.split(delimiter)
                    two_columns.extend([(split_line[0], ' '.join(split_line[1:]))])

                with open(output_file, 'w') as f_out:
                    for row in two_columns:
                        f_out.write('\t'.join(row) + '\n')
                file_counter += 1
    for filename in os.listdir(nucleotide_folder_path):
        if filename.endswith('.txt'):  # Process only .txt files
            input_file2 = os.path.join(nucleotide_folder_path, filename)
            output_file2 = os.path.join(nucleotide_folder_path, f"processed_{filename}")

            # Read the contents of the .txt file
            with open(input_file2, 'r') as txt_file:
                lines = txt_file.readlines()

            # Apply formula to multiply values by 10 (if the value is not 0)
            modified_lines = []
            for line in lines:
                line_data = line.strip().split('\t')  # Assuming tab-separated data, modify delimiter if needed
                if len(line_data) >= 2:
                    try:
                        value = float(line_data[1])
                        if value != 0:
                            line_data.append(str(value * 10))
                        else:
                            line_data.append('0')
                    except ValueError:
                        line_data.append('Invalid')  # Or handle the invalid value case differently

                modified_lines.append('\t'.join(line_data))

            # Write the modified data to a new output file
            with open(output_file2, 'w') as output_txt_file:
                output_txt_file.write('\n'.join(modified_lines))


def compile_distance_rmsd(path_for_compiled_data, F_or_G, mutant, interprotomer_folder_path, nucleotide_folder_path,
                          chains, target_chain_number):
    # first prepare excel sheets
    if F_or_G == 'Factin':
        max_column = 14
        # file offsets are used for calculating columns to paste data into.
        # this is just the file number (aka WT file numbering starts at 5)
        # since dimer/trimer have different numbers of files, it allows flexibility
        interprotomer_chain1 = chains[target_chain_number - 2]
        interprotomer_chain2 = chains[target_chain_number + 2]
        target_chain = chains[target_chain_number]
        mutant_file_offset = 1
        WT_file_offset = 3
        interprotomer_headers = [
            f'{mutant}_{F_or_G}_Chain_{interprotomer_chain1}_{target_chain}',
            f'{mutant}_{F_or_G}_Chain_{target_chain}_{interprotomer_chain2}',
            f'WT_{F_or_G}_Chain_{interprotomer_chain1}_{target_chain}',
            f'WT_{F_or_G}_Chain_{target_chain}_{interprotomer_chain2}'
        ]

        nucleotide_headers = [
            f'{mutant}_{F_or_G}_Chain_{target_chain}_ADP',
            f'{mutant}_{F_or_G}_Chain_{target_chain}_PO4',
            f'WT_{F_or_G}_Chain_{target_chain}_ADP',
            f'WT_{F_or_G}_Chain_{target_chain}_PO4'
        ]

    if F_or_G == 'Gactin':
        max_column = 14
        mutant_file_offset = 1
        WT_file_offset = 3
        nucleotide_headers = [
            f'{mutant}_{F_or_G}_protein_ADP',
            f'{mutant}_{F_or_G}_protein_PO4',
            f'WT_{F_or_G}_protein_ADP',
            f'WT_{F_or_G}_protein_PO4'
        ]

    wb = load_workbook(f'{path_for_compiled_data}{mutant}_{F_or_G}_COMPILED.xlsx')
    ws1 = wb[f'{mutant}_RMSD']

    if 'Replicate1' in interprotomer_folder_path:
        # Create a new worksheet
        ws8 = wb.create_sheet(title='Interprotomer_rmsd')
        ws9 = wb.create_sheet(title='Nucleotide_rmsd')
    else:
        ws8 = wb[f'Interprotomer_rmsd']
        ws9 = wb[f'Nucleotide_rmsd']

    # copy time from rmsd worksheet
    for row_index in range(1, ws1.max_row + 1):
        # Get the value from the source column and assign it to the destination column
        ws8.cell(row=row_index + 1, column=1).value = ws1.cell(row=row_index, column=1).value
        ws8.cell(row=row_index + 1, column=2).value = ws1.cell(row=row_index, column=2).value
        ws9.cell(row=row_index + 1, column=1).value = ws1.cell(row=row_index, column=1).value
        ws9.cell(row=row_index + 1, column=2).value = ws1.cell(row=row_index, column=2).value

    col = 3
    if F_or_G == "Factin":
        for interprotomer_header in interprotomer_headers:
            cell = ws8.cell(row=1, column=col)
            cell.value = interprotomer_header
            col += 6

    col = 3
    for nucleotide_header in nucleotide_headers:
        cell = ws9.cell(row=1, column=col)
        cell.value = nucleotide_header
        col += 6
    #merge header cells
    for col in range(3, max_column, 6):
        merge_range = f'{ws8.cell(row=1, column=col).coordinate}:{ws8.cell(row=1, column=col + 5).coordinate}'
        ws8.merge_cells(merge_range)
        merge_range = f'{ws9.cell(row=1, column=col).coordinate}:{ws9.cell(row=1, column=col + 5).coordinate}'
        ws9.merge_cells(merge_range)

    # Assign alternating values "mutant" and "WT" to individual cells before merging
    for col in range(3, max_column, 6):
        cell = ws8.cell(row=2, column=col)
        cell.value = f'{mutant}'
        cell = ws9.cell(row=2, column=col)
        cell.value = f'{mutant}'
    for col in range(6, max_column, 6):
        cell = ws8.cell(row=2, column=col)
        cell.value = 'WT'
        cell = ws9.cell(row=2, column=col)
        cell.value = 'WT'

    # Merge cells after assigning values
    for col in range(3, max_column, 3):
        merge_range = f'{ws8.cell(row=2, column=col).coordinate}:{ws8.cell(row=2, column=col + 2).coordinate}'
        ws8.merge_cells(merge_range)
        merge_range = f'{ws9.cell(row=2, column=col).coordinate}:{ws9.cell(row=2, column=col + 2).coordinate}'
        ws9.merge_cells(merge_range)

    # Assign "Replicate2" to even cells (if col value divided by 2 is true aka 0), otherwise "Replicate1"
    for col in range(3, max_column + 2, 3):
        cell = ws8.cell(row=3, column=col)
        cell.value = "Replicate1"
        cell = ws9.cell(row=3, column=col)
        cell.value = "Replicate1"
        col += 1
        cell = ws8.cell(row=3, column=col)
        cell.value = "Replicate2"
        cell = ws9.cell(row=3, column=col)
        cell.value = "Replicate2"
        col += 1
        cell = ws8.cell(row=3, column=col)
        cell.value = "Replicate3"
        cell = ws9.cell(row=3, column=col)
        cell.value = "Replicate3"
        col += 1
        # Iterate through each .txt file in the directory

    if 'Replicate1' in interprotomer_folder_path:
        for file_name in sorted(os.listdir(interprotomer_folder_path)):
            if file_name.startswith('processed_') and file_name.endswith('.txt'):
                file_number = int(file_name.split('_')[1].split('.')[0])  # Extract the file number

                # Read the contents of column B from the .txt file
                with open(os.path.join(interprotomer_folder_path, file_name), 'r') as file:
                    column_C_data = [line.split()[2] for line in file.readlines()]

                if f'{mutant}' in file_name:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column = 3 + 6 * (file_number - mutant_file_offset)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws8.cell(row=i + 3, column=paste_column, value=value)
                if 'WT' in file_name:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column = 6 + 6 * (file_number - WT_file_offset)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws8.cell(row=i + 3, column=paste_column, value=value)
    if 'Replicate2' in interprotomer_folder_path:
        for file_name in sorted(os.listdir(interprotomer_folder_path)):
            if file_name.startswith('processed_') and file_name.endswith('.txt'):
                file_number = int(file_name.split('_')[1].split('.')[0])  # Extract the file number

                # Read the contents of column B from the .txt file
                with open(os.path.join(interprotomer_folder_path, file_name), 'r') as file:
                    column_C_data = [line.split()[2] for line in file.readlines()]

                if f'{mutant}' in file_name:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column = 4 + 6 * (file_number - mutant_file_offset)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws8.cell(row=i + 3, column=paste_column, value=value)
                if 'WT' in file_name:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column = 7 + 6 * (file_number - WT_file_offset)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws8.cell(row=i + 3, column=paste_column, value=value)
    if 'Replicate3' in interprotomer_folder_path:
        for file_name in sorted(os.listdir(interprotomer_folder_path)):
            if file_name.startswith('processed_') and file_name.endswith('.txt'):
                file_number = int(file_name.split('_')[1].split('.')[0])  # Extract the file number

                # Read the contents of column B from the .txt file
                with open(os.path.join(interprotomer_folder_path, file_name), 'r') as file:
                    column_C_data = [line.split()[2] for line in file.readlines()]

                if f'{mutant}' in file_name:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column = 5 + 6 * (file_number - mutant_file_offset)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws8.cell(row=i + 3, column=paste_column, value=value)
                if 'WT' in file_name:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column = 8 + 6 * (file_number - WT_file_offset)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws8.cell(row=i + 3, column=paste_column, value=value)

    if 'Replicate1' in nucleotide_folder_path:
        for file_name in sorted(os.listdir(nucleotide_folder_path)):
            if file_name.startswith('processed_') and file_name.endswith('.txt'):
                file_number = int(file_name.split('_')[1].split('.')[0])  # Extract the file number

                # Read the contents of column B from the .txt file
                with open(os.path.join(nucleotide_folder_path, file_name), 'r') as file:
                    column_C_data = [line.split()[2] for line in file.readlines()]

                if f'{mutant}' in file_name:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column = 3 + 6 * (file_number - mutant_file_offset)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws9.cell(row=i + 3, column=paste_column, value=value)
                if 'WT' in file_name:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column = 6 + 6 * (file_number - WT_file_offset)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws9.cell(row=i + 3, column=paste_column, value=value)
    if 'Replicate2' in nucleotide_folder_path:
        for file_name in sorted(os.listdir(nucleotide_folder_path)):
            if file_name.startswith('processed_') and file_name.endswith('.txt'):
                file_number = int(file_name.split('_')[1].split('.')[0])  # Extract the file number

                # Read the contents of column B from the .txt file
                with open(os.path.join(nucleotide_folder_path, file_name), 'r') as file:
                    column_C_data = [line.split()[2] for line in file.readlines()]

                if f'{mutant}' in file_name:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column = 4 + 6 * (file_number - mutant_file_offset)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws9.cell(row=i + 3, column=paste_column, value=value)
                if 'WT' in file_name:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column = 7 + 6 * (file_number - WT_file_offset)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws9.cell(row=i + 3, column=paste_column, value=value)

    if 'Replicate3' in nucleotide_folder_path:
        for file_name in sorted(os.listdir(nucleotide_folder_path)):
            if file_name.startswith('processed_') and file_name.endswith('.txt'):
                file_number = int(file_name.split('_')[1].split('.')[0])  # Extract the file number

                # Read the contents of column B from the .txt file
                with open(os.path.join(nucleotide_folder_path, file_name), 'r') as file:
                    column_C_data = [line.split()[2] for line in file.readlines()]

                if f'{mutant}' in file_name:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column = 5 + 6 * (file_number - mutant_file_offset)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws9.cell(row=i + 3, column=paste_column, value=value)
                if 'WT' in file_name:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column = 8 + 6 * (file_number - WT_file_offset)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws9.cell(row=i + 3, column=paste_column, value=value)

    wb.save(f'{path_for_compiled_data}{mutant}_{F_or_G}_COMPILED.xlsx')


def process_distance_files(path_for_compiled_data, F_or_G, mutant, number_of_timepoints, allostery_folder_path,
                           interprotomer_folder_path, interstrand_folder_path, nucleotide_folder_path,
                           chains, target_chain_number, nucleotide, nucleotide_state):
    if "Replicate1" in allostery_folder_path and F_or_G == 'Factin':
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Allostery'
        ws2 = wb.create_sheet(title='Nucleotide')
        ws3 = wb.create_sheet(title='Interstrand')
        ws4 = wb.create_sheet(title='Interprotomer')
        
        

    if "Replicate1" in allostery_folder_path and F_or_G == 'Gactin':
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Allostery'
        ws2 = wb.create_sheet(title='Nucleotide')

    #max column is based on whatever sheet will need the most columns
    #here, interprotomer_naming is the longest. It's 20 rows * 3 replicates * 2 (mutant, WT) = 120
    #it's needed for determining where to stop merging header cells
    max_column = 120
    target_chain = chains[target_chain_number]
    col = 3
    file_counter = 1
    delimiter = ' '
    
    allostery_naming = [
        f'Chain_{target_chain}_E107_Chain_{target_chain}_Q137',
        f'Chain_{target_chain}_F375_Chain_{target_chain}_R116',
        f'Chain_{target_chain}_H88_Chain_{target_chain}_D56',
        f'Chain_{target_chain}_H88_Chain_{target_chain}_V54',
        f'Chain_{target_chain}_H371_Chain_{target_chain}_E117',
        f'Chain_{target_chain}_I76_Chain_{target_chain}_W79',
        f'Chain_{target_chain}_I122_Chain_{target_chain}_W86',
        f'Chain_{target_chain}_K118_Chain_{target_chain}_W79',
        f'Chain_{target_chain}_M119_Chain_{target_chain}_W79',
        f'Chain_{target_chain}_M123_Chain_{target_chain}_W86',
        f'Chain_{target_chain}_N111_Chain_{target_chain}_G74',
        f'Chain_{target_chain}_N115_Chain_{target_chain}_I76',
        f'Chain_{target_chain}_R116_Chain_{target_chain}_E107'
    ]
        
    # add headers to excel file
    if "Replicate1" in allostery_folder_path:
        for allostery_header in allostery_naming:
            cell = ws1.cell(row=1, column=col)
            cell.value = allostery_header
            col += 6

    # first, open each .xvg file, strip non-data lines and save it in new format
    # run for PBM files first
    for allostery_name in allostery_naming:
        for filename in os.listdir(allostery_folder_path):
            if mutant in filename and allostery_name in filename and filename.endswith('.xvg'):
                input_file = os.path.join(allostery_folder_path, filename)
                output_file = os.path.join(allostery_folder_path,
                                           f"{file_counter}_{mutant}_{F_or_G}_{allostery_name}.txt")
                with open(input_file, 'r') as f:
                    lines = f.readlines()
                # Remove lines starting with '#' or '@' (common in .xvg files)
                lines = [line.strip() for line in lines if
                         not line.startswith('#') and not line.startswith('@') and line.strip()]
                two_columns = []
                for line in lines:
                    split_line = line.split(delimiter)
                    two_columns.extend([(split_line[0], ' '.join(split_line[1:]))])
                with open(output_file, 'w') as f_out:
                    for row in two_columns:
                        f_out.write('\t'.join(row) + '\n')
                file_counter += 1
    # next run the same commands for WT files
    for allostery_name in allostery_naming:
        for filename in os.listdir(allostery_folder_path):
            if 'WT' in filename and allostery_name in filename and filename.endswith('.xvg'):
                input_file = os.path.join(allostery_folder_path, filename)
                output_file = os.path.join(allostery_folder_path,
                                           f"{file_counter}_WT_{F_or_G}_{allostery_name}.txt")
                with open(input_file, 'r') as f:
                    lines = f.readlines()
                # Remove lines starting with '#' or '@' (common in .xvg files)
                lines = [line.strip() for line in lines if
                         not line.startswith('#') and not line.startswith('@') and line.strip()]
                two_columns = []
                for line in lines:
                    split_line = line.split(delimiter)
                    two_columns.extend([(split_line[0], ' '.join(split_line[1:]))])
                with open(output_file, 'w') as f_out:
                    for row in two_columns:
                        f_out.write('\t'.join(row) + '\n')
                file_counter += 1

    for allostery_name in allostery_naming:
        for filename in os.listdir(allostery_folder_path):
            if allostery_name in filename and filename.endswith('.txt'):
                input_file = os.path.join(allostery_folder_path, filename)
                output_file = os.path.join(allostery_folder_path, f"processed_{filename}")
                # Read the contents of the .txt file
                with open(input_file, 'r') as txt_file:
                    lines = txt_file.readlines()

                # Apply formula to multiply values by 10 (if the value is not 0)
                modified_lines = []
                for line in lines:
                    line_data = line.strip().split('\t')  # Assuming tab-separated data, modify delimiter if needed
                    if len(line_data) >= 2:
                        try:
                            value = float(line_data[1])
                            if value != 0:
                                line_data.append(str(value * 10))
                            else:
                                line_data.append('0')
                        except ValueError:
                            line_data.append('Invalid')  # Or handle the invalid value case differently

                    modified_lines.append('\t'.join(line_data))

                # Write the modified data to a new output file
                with open(output_file, 'w') as output_txt_file:
                    output_txt_file.write('\n'.join(modified_lines))

    if F_or_G == 'Factin':
        col = 3
        file_counter = 1
        interprotomer_chain1 = chains[target_chain_number - 2]
        interprotomer_chain2 = chains[target_chain_number + 2]
        interprotomer_naming = [
            f'Chain_{interprotomer_chain1}_A170_Chain_{target_chain}_Q41',
            f'Chain_{interprotomer_chain1}_E167_Chain_{target_chain}_Q49',
            f'Chain_{interprotomer_chain1}_F352_Chain_{target_chain}_V45',
            f'Chain_{interprotomer_chain1}_F375_Chain_{target_chain}_Q41',
            f'Chain_{interprotomer_chain1}_F375_Chain_{target_chain}_V43',
            f'Chain_{interprotomer_chain1}_L171_Chain_{target_chain}_H40',
            f'Chain_{interprotomer_chain1}_T351_Chain_{target_chain}_V45',
            f'Chain_{interprotomer_chain1}_Y169_Chain_{target_chain}_L50',
            f'Chain_{interprotomer_chain1}_Y169_Chain_{target_chain}_P38',
            f'Chain_{interprotomer_chain1}_Y169_Chain_{target_chain}_Q49',
            f'Chain_{target_chain}_A170_Chain_{interprotomer_chain2}_Q41',
            f'Chain_{target_chain}_E167_Chain_{interprotomer_chain2}_Q49',
            f'Chain_{target_chain}_F352_Chain_{interprotomer_chain2}_V45',
            f'Chain_{target_chain}_F375_Chain_{interprotomer_chain2}_Q41',
            f'Chain_{target_chain}_F375_Chain_{interprotomer_chain2}_V43',
            f'Chain_{target_chain}_L171_Chain_{interprotomer_chain2}_H40',
            f'Chain_{target_chain}_T351_Chain_{interprotomer_chain2}_V45',
            f'Chain_{target_chain}_Y169_Chain_{interprotomer_chain2}_L50',
            f'Chain_{target_chain}_Y169_Chain_{interprotomer_chain2}_P38',
            f'Chain_{target_chain}_Y169_Chain_{interprotomer_chain2}_Q49'
        ]
        # add headers to excel file
        if "Replicate1" in interprotomer_folder_path:
            for interprotomer_header in interprotomer_naming:
                cell = ws4.cell(row=1, column=col)
                cell.value = interprotomer_header
                col += 6

        # first, open each .xvg file, strip non-data lines and save it in new format
        # run for PBM files first
        for interprotomer_name in interprotomer_naming:
            for filename in os.listdir(interprotomer_folder_path):
                if mutant in filename and interprotomer_name in filename and filename.endswith('.xvg'):
                    input_file = os.path.join(interprotomer_folder_path, filename)
                    output_file = os.path.join(interprotomer_folder_path,
                                               f"{file_counter}_{mutant}_{F_or_G}_{interprotomer_name}.txt")
                    with open(input_file, 'r') as f:
                        lines = f.readlines()
                    # Remove lines starting with '#' or '@' (common in .xvg files)
                    lines = [line.strip() for line in lines if
                             not line.startswith('#') and not line.startswith('@') and line.strip()]
                    two_columns = []
                    for line in lines:
                        split_line = line.split(delimiter)
                        two_columns.extend([(split_line[0], ' '.join(split_line[1:]))])
                    with open(output_file, 'w') as f_out:
                        for row in two_columns:
                            f_out.write('\t'.join(row) + '\n')
                    file_counter += 1
        # next run the same commands for WT files
        for interprotomer_name in interprotomer_naming:
            for filename in os.listdir(interprotomer_folder_path):
                if 'WT' in filename and interprotomer_name in filename and filename.endswith('.xvg'):
                    input_file = os.path.join(interprotomer_folder_path, filename)
                    output_file = os.path.join(interprotomer_folder_path,
                                               f"{file_counter}_WT_{F_or_G}_{interprotomer_name}.txt")
                    with open(input_file, 'r') as f:
                        lines = f.readlines()
                    # Remove lines starting with '#' or '@' (common in .xvg files)
                    lines = [line.strip() for line in lines if
                             not line.startswith('#') and not line.startswith('@') and line.strip()]
                    two_columns = []
                    for line in lines:
                        split_line = line.split(delimiter)
                        two_columns.extend([(split_line[0], ' '.join(split_line[1:]))])
                    with open(output_file, 'w') as f_out:
                        for row in two_columns:
                            f_out.write('\t'.join(row) + '\n')
                    file_counter += 1

        for interprotomer_name in interprotomer_naming:
            for filename in os.listdir(interprotomer_folder_path):
                if interprotomer_name in filename and filename.endswith('.txt'):
                    input_file = os.path.join(interprotomer_folder_path, filename)
                    output_file = os.path.join(interprotomer_folder_path, f"processed_{filename}")
                    # Read the contents of the .txt file
                    with open(input_file, 'r') as txt_file:
                        lines = txt_file.readlines()

                    # Apply formula to multiply values by 10 (if the value is not 0)
                    modified_lines = []
                    for line in lines:
                        line_data = line.strip().split('\t')  # Assuming tab-separated data, modify delimiter if needed
                        if len(line_data) >= 2:
                            try:
                                value = float(line_data[1])
                                if value != 0:
                                    line_data.append(str(value * 10))
                                else:
                                    line_data.append('0')
                            except ValueError:
                                line_data.append('Invalid')  # Or handle the invalid value case differently

                        modified_lines.append('\t'.join(line_data))

                    # Write the modified data to a new output file
                    with open(output_file, 'w') as output_txt_file:
                        output_txt_file.write('\n'.join(modified_lines))

    if F_or_G == 'Factin':
        col = 3
        file_counter = 1
        #note: because 'chains' is an index with multiple entries, numbering starts at 0 instead of 1
        #so numbering needs to be adjusted accordingly
        interstrand_chain1 = chains[target_chain_number - 1]
        interstrand_chain2 = chains[target_chain_number + 1]
        interstrand_naming = [
            f'Chain_{interstrand_chain1}_E270_Chain_{target_chain}_R39',
            f'Chain_{interstrand_chain1}_E270_Chain_{target_chain}_T66',
            f'Chain_{interstrand_chain1}_E270_Chain_{target_chain}_T202',
            f'Chain_{interstrand_chain1}_E270_Chain_{target_chain}_T203',
            f'Chain_{interstrand_chain1}_G268_Chain_{target_chain}_R39',
            f'Chain_{target_chain}_H173_Chain_{interstrand_chain2}_G268',
            f'Chain_{target_chain}_H173_Chain_{interstrand_chain2}_I267',
            f'Chain_{target_chain}_K113_Chain_{interstrand_chain2}_E195',
            f'Chain_{target_chain}_P112_Chain_{interstrand_chain2}_T194'
        ]

        if "Replicate1" in interstrand_folder_path:
            for interstrand_header in interstrand_naming:
                cell = ws3.cell(row=1, column=col)
                cell.value = interstrand_header
                col += 6

        # first, open each .xvg file, strip non-data lines and save it in new format
        # run for PBM files first
        for interstrand_name in interstrand_naming:
            for filename in os.listdir(interstrand_folder_path):
                if mutant in filename and interstrand_name in filename and filename.endswith('.xvg'):
                    input_file = os.path.join(interstrand_folder_path, filename)
                    output_file = os.path.join(interstrand_folder_path,
                                               f"{file_counter}_{mutant}_{F_or_G}_{interstrand_name}.txt")
                    with open(input_file, 'r') as f:
                        lines = f.readlines()
                    # Remove lines starting with '#' or '@' (common in .xvg files)
                    lines = [line.strip() for line in lines if
                             not line.startswith('#') and not line.startswith('@') and line.strip()]
                    two_columns = []
                    for line in lines:
                        split_line = line.split(delimiter)
                        two_columns.extend([(split_line[0], ' '.join(split_line[1:]))])
                    with open(output_file, 'w') as f_out:
                        for row in two_columns:
                            f_out.write('\t'.join(row) + '\n')
                    file_counter += 1
        # next run the same commands for WT files
        for interstrand_name in interstrand_naming:
            for filename in os.listdir(interstrand_folder_path):
                if 'WT' in filename and interstrand_name in filename and filename.endswith('.xvg'):
                    input_file = os.path.join(interstrand_folder_path, filename)
                    output_file = os.path.join(interstrand_folder_path,
                                               f"{file_counter}_WT_{F_or_G}_{interstrand_name}.txt")
                    with open(input_file, 'r') as f:
                        lines = f.readlines()
                    # Remove lines starting with '#' or '@' (common in .xvg files)
                    lines = [line.strip() for line in lines if
                             not line.startswith('#') and not line.startswith('@') and line.strip()]
                    two_columns = []
                    for line in lines:
                        split_line = line.split(delimiter)
                        two_columns.extend([(split_line[0], ' '.join(split_line[1:]))])
                    with open(output_file, 'w') as f_out:
                        for row in two_columns:
                            f_out.write('\t'.join(row) + '\n')
                    file_counter += 1

        for interstrand_name in interstrand_naming:
            for filename in os.listdir(interstrand_folder_path):
                if interstrand_name in filename and filename.endswith('.txt'):
                    input_file = os.path.join(interstrand_folder_path, filename)
                    output_file = os.path.join(interstrand_folder_path, f"processed_{filename}")
                    # Read the contents of the .txt file
                    with open(input_file, 'r') as txt_file:
                        lines = txt_file.readlines()

                    # Apply formula to multiply values by 10 (if the value is not 0)
                    modified_lines = []
                    for line in lines:
                        line_data = line.strip().split('\t')  # Assuming tab-separated data, modify delimiter if needed
                        if len(line_data) >= 2:
                            try:
                                value = float(line_data[1])
                                if value != 0:
                                    line_data.append(str(value * 10))
                                else:
                                    line_data.append('0')
                            except ValueError:
                                line_data.append('Invalid')  # Or handle the invalid value case differently

                        modified_lines.append('\t'.join(line_data))

                    # Write the modified data to a new output file
                    with open(output_file, 'w') as output_txt_file:
                        output_txt_file.write('\n'.join(modified_lines))

    col = 3
    file_counter = 1
    if nucleotide_state == 'ADP-Pi':
        nucleotide_naming = [
            f'Chain_{target_chain}_ADP_Chain_{target_chain}_D157',
            f'Chain_{target_chain}_ADP_Chain_{target_chain}_E214',
            f'Chain_{target_chain}_ADP_Chain_{target_chain}_K18',
            f'Chain_{target_chain}_ADP_Chain_{target_chain}_K213',
            f'Chain_{target_chain}_ADP_Chain_{target_chain}_K336',
            f'Chain_{target_chain}_ADP_Chain_{target_chain}_L16',
            f'Chain_{target_chain}_ADP_Chain_{target_chain}_M305',
            f'Chain_{target_chain}_ADP_Chain_{target_chain}_R210',
            f'Chain_{target_chain}_ADP_Chain_{target_chain}_T303',
            f'Chain_{target_chain}_ADP_Chain_{target_chain}_Y306',
            f'Chain_{target_chain}_PO4_Chain_{target_chain}_D157',
            f'Chain_{target_chain}_PO4_Chain_{target_chain}_Q137',
            f'Chain_{target_chain}_PO4_Chain_{target_chain}_S14',
            f'Chain_{target_chain}_PO4_Chain_{target_chain}_V159',
            f'Chain_{target_chain}_S14_Chain_{target_chain}_G74'
        ]
    else:
        nucleotide_naming = [
            f'Chain_{target_chain}_{nucleotide}_Chain_{target_chain}_D157',
            f'Chain_{target_chain}_{nucleotide}_Chain_{target_chain}_E214',
            f'Chain_{target_chain}_{nucleotide}_Chain_{target_chain}_K18',
            f'Chain_{target_chain}_{nucleotide}_Chain_{target_chain}_K213',
            f'Chain_{target_chain}_{nucleotide}_Chain_{target_chain}_K336',
            f'Chain_{target_chain}_{nucleotide}_Chain_{target_chain}_L16',
            f'Chain_{target_chain}_{nucleotide}_Chain_{target_chain}_M305',
            f'Chain_{target_chain}_{nucleotide}_Chain_{target_chain}_R210',
            f'Chain_{target_chain}_{nucleotide}_Chain_{target_chain}_T303',
            f'Chain_{target_chain}_{nucleotide}_Chain_{target_chain}_Y306',
            f'Chain_{target_chain}_S14_Chain_{target_chain}_G74'
        ]
    # add headers to excel file
    if "Replicate1" in nucleotide_folder_path and F_or_G == 'Factin':
        for nucleotide_header in nucleotide_naming:
            cell = ws2.cell(row=1, column=col)
            cell.value = nucleotide_header
            col += 6
    if "Replicate1" in allostery_folder_path and F_or_G == 'Gactin':
        for nucleotide_header in nucleotide_naming:
            cell = ws2.cell(row=1, column=col)
            cell.value = nucleotide_header
            col += 6

    # first, open each .xvg file, strip non-data lines and save it in new format
    # run for PBM files first
    for nucleotide_name in nucleotide_naming:
        for filename in os.listdir(nucleotide_folder_path):
            if mutant in filename and nucleotide_name in filename and filename.endswith('.xvg'):
                input_file = os.path.join(nucleotide_folder_path, filename)
                output_file = os.path.join(nucleotide_folder_path,
                                           f"{file_counter}_{mutant}_{F_or_G}_{nucleotide_name}.txt")
                with open(input_file, 'r') as f:
                    lines = f.readlines()
                # Remove lines starting with '#' or '@' (common in .xvg files)
                lines = [line.strip() for line in lines if
                         not line.startswith('#') and not line.startswith('@') and line.strip()]
                two_columns = []
                for line in lines:
                    split_line = line.split(delimiter)
                    two_columns.extend([(split_line[0], ' '.join(split_line[1:]))])
                with open(output_file, 'w') as f_out:
                    for row in two_columns:
                        f_out.write('\t'.join(row) + '\n')
                file_counter += 1
    # next run the same commands for WT files
    for nucleotide_name in nucleotide_naming:
        for filename in os.listdir(nucleotide_folder_path):
            if 'WT' in filename and nucleotide_name in filename and filename.endswith('.xvg'):
                input_file = os.path.join(nucleotide_folder_path, filename)
                output_file = os.path.join(nucleotide_folder_path,
                                           f"{file_counter}_WT_{F_or_G}_{nucleotide_name}.txt")
                with open(input_file, 'r') as f:
                    lines = f.readlines()
                # Remove lines starting with '#' or '@' (common in .xvg files)
                lines = [line.strip() for line in lines if
                         not line.startswith('#') and not line.startswith('@') and line.strip()]
                two_columns = []
                for line in lines:
                    split_line = line.split(delimiter)
                    two_columns.extend([(split_line[0], ' '.join(split_line[1:]))])
                with open(output_file, 'w') as f_out:
                    for row in two_columns:
                        f_out.write('\t'.join(row) + '\n')
                file_counter += 1

    for nucleotide_name in nucleotide_naming:
        for filename in os.listdir(nucleotide_folder_path):
            if nucleotide_name in filename and filename.endswith('.txt'):
                input_file = os.path.join(nucleotide_folder_path, filename)
                output_file = os.path.join(nucleotide_folder_path, f"processed_{filename}")
                # Read the contents of the .txt file
                with open(input_file, 'r') as txt_file:
                    lines = txt_file.readlines()

                # Apply formula to multiply values by 10 (if the value is not 0)
                modified_lines = []
                for line in lines:
                    line_data = line.strip().split('\t')  # Assuming tab-separated data, modify delimiter if needed
                    if len(line_data) >= 2:
                        try:
                            value = float(line_data[1])
                            if value != 0:
                                line_data.append(str(value * 10))
                            else:
                                line_data.append('0')
                        except ValueError:
                            line_data.append('Invalid')  # Or handle the invalid value case differently

                    modified_lines.append('\t'.join(line_data))

                # Write the modified data to a new output file
                with open(output_file, 'w') as output_txt_file:
                    output_txt_file.write('\n'.join(modified_lines))

    if "Replicate1" in allostery_folder_path:
        # merge header cells
        for col in range(3, max_column, 6):
            merge_range = f'{ws1.cell(row=1, column=col).coordinate}:{ws1.cell(row=1, column=col + 5).coordinate}'
            ws1.merge_cells(merge_range)
            merge_range = f'{ws2.cell(row=1, column=col).coordinate}:{ws2.cell(row=1, column=col + 5).coordinate}'
            ws2.merge_cells(merge_range)
            if F_or_G == 'Factin':
                merge_range = f'{ws3.cell(row=1, column=col).coordinate}:{ws3.cell(row=1, column=col + 5).coordinate}'
                ws3.merge_cells(merge_range)
                merge_range = f'{ws4.cell(row=1, column=col).coordinate}:{ws4.cell(row=1, column=col + 5).coordinate}'
                ws4.merge_cells(merge_range)

        # Assign alternating values for PBM and "WT" to individual cells
        for col in range(3, max_column, 6):
            cell = ws1.cell(row=2, column=col)
            cell.value = f'{mutant}'
            cell = ws2.cell(row=2, column=col)
            cell.value = f'{mutant}'
            if F_or_G == 'Factin':
                cell = ws3.cell(row=2, column=col)
                cell.value = f'{mutant}'
                cell = ws4.cell(row=2, column=col)
                cell.value = f'{mutant}'
        for col in range(6, max_column, 6):
            cell = ws1.cell(row=2, column=col)
            cell.value = 'WT'
            cell = ws2.cell(row=2, column=col)
            cell.value = 'WT'
            if F_or_G == 'Factin':
                cell = ws3.cell(row=2, column=col)
                cell.value = 'WT'
                cell = ws4.cell(row=2, column=col)
                cell.value = 'WT'

        # merge PBM and WT header cells
        for col in range(3, max_column, 6):
            merge_range = f'{ws1.cell(row=2, column=col).coordinate}:{ws1.cell(row=2, column=col + 2).coordinate}'
            ws1.merge_cells(merge_range)
            merge_range = f'{ws2.cell(row=2, column=col).coordinate}:{ws2.cell(row=2, column=col + 2).coordinate}'
            ws2.merge_cells(merge_range)
            if F_or_G == 'Factin':
                merge_range = f'{ws3.cell(row=2, column=col).coordinate}:{ws3.cell(row=2, column=col + 2).coordinate}'
                ws3.merge_cells(merge_range)
                merge_range = f'{ws4.cell(row=2, column=col).coordinate}:{ws4.cell(row=2, column=col + 2).coordinate}'
                ws4.merge_cells(merge_range)

        # add Replicate headers
        for col in range(3, max_column + 1, 3):
            cell = ws1.cell(row=5, column=col)
            cell.value = "Replicate1"
            cell = ws2.cell(row=5, column=col)
            cell.value = "Replicate1"
            if F_or_G == 'Factin':
                cell = ws3.cell(row=5, column=col)
                cell.value = "Replicate1"
                cell = ws4.cell(row=5, column=col)
                cell.value = "Replicate1"
            col += 1
            cell = ws1.cell(row=5, column=col)
            cell.value = "Replicate2"
            cell = ws2.cell(row=5, column=col)
            cell.value = "Replicate2"
            if F_or_G == 'Factin':
                cell = ws3.cell(row=5, column=col)
                cell.value = "Replicate2"
                cell = ws4.cell(row=5, column=col)
                cell.value = "Replicate2"
            col += 1
            cell = ws1.cell(row=5, column=col)
            cell.value = "Replicate3"
            cell = ws2.cell(row=5, column=col)
            cell.value = "Replicate3"
            if F_or_G == 'Factin':
                cell = ws3.cell(row=5, column=col)
                cell.value = "Replicate3"
                cell = ws4.cell(row=5, column=col)
                cell.value = "Replicate3"
            col += 1

        # add formula for converting time to ns
        for row in range(6, 7 + number_of_timepoints):
            formula = f'=A{row}/1000'  # Change the formula as needed
            ws1[f'B{row}'].value = formula
            ws2[f'B{row}'].value = formula
            if F_or_G == 'Factin':
                ws3[f'B{row}'].value = formula
                ws4[f'B{row}'].value = formula

        # add headers for formulas
        ws1['B3'] = 'Average Distance'
        ws2['B3'] = 'Average Distance'
        ws1['B4'] = 'Occupancy %'
        ws2['B4'] = 'Occupancy %'
        if F_or_G == 'Factin':
            ws3['B3'] = 'Average Distance'
            ws4['B3'] = 'Average Distance'
            ws3['B4'] = 'Occupancy %'
            ws4['B4'] = 'Occupancy %'

        # add formulas to rows 3/4 to calculate average distance, and determine occupancy percentage
        for col in range(3, max_column + 1, 3):
            col_letter = get_column_letter(col)
            next_col_letter = get_column_letter(col + 2)

            formula1 = f'=AVERAGE({col_letter}6:{next_col_letter}{number_of_timepoints + 6})'
            formula2 = f'=(COUNTIF({col_letter}6:{next_col_letter}{number_of_timepoints + 6}, "<=3.5") / COUNT({col_letter}6:{next_col_letter}{number_of_timepoints + 6}))'
            ws1[f"{col_letter}3"].value = formula1
            ws1[f"{col_letter}4"].value = formula2
            ws2[f"{col_letter}3"].value = formula1
            ws2[f"{col_letter}4"].value = formula2
            if F_or_G == 'Factin':
                ws3[f"{col_letter}3"].value = formula1
                ws3[f"{col_letter}4"].value = formula2
                ws4[f"{col_letter}3"].value = formula1
                ws4[f"{col_letter}4"].value = formula2

        wb.save(f'{path_for_compiled_data}{mutant}_{F_or_G}_DISTANCES.xlsx')


def compile_distances(F_or_G, mutant, path_for_compiled_data, nucleotide_state,
                      allostery_folder_path_rep1, allostery_folder_path_rep2, allostery_folder_path_rep3,
                      interprotomer_folder_path_rep1, interprotomer_folder_path_rep2, interprotomer_folder_path_rep3,
                      interstrand_folder_path_rep1, interstrand_folder_path_rep2, interstrand_folder_path_rep3,
                      nucleotide_folder_path_rep1, nucleotide_folder_path_rep2, nucleotide_folder_path_rep3):
    
    
    if F_or_G == 'Factin':
        # offset number is the start number of the WT files, used for calculating what column to paste into
        interprotomer_filenumber_offset = 21
        interstrand_filenumber_offset = 10
        
    allostery_filenumber_offset = 14
    if nucleotide_state == 'ADP-Pi':
        nucleotide_filenumber_offset = 16
    if nucleotide_state == 'ADP':
        nucleotide_filenumber_offset = 12
    if nucleotide_state == 'ATP':
        nucleotide_filenumber_offset = 12


    wb = load_workbook(f'{path_for_compiled_data}{mutant}_{F_or_G}_DISTANCES.xlsx')
    if F_or_G == 'Factin':
        ws1 = wb['Allostery']
        ws2 = wb['Nucleotide']
        ws3 = wb['Interstrand']
        ws4 = wb['Interprotomer']
        
        
    
    if F_or_G == 'Gactin':
        ws1 = wb['Allostery']
        ws2 = wb['Nucleotide']

    for filename in sorted(os.listdir(allostery_folder_path_rep1)):
        if filename.startswith('processed_') and filename.endswith('.txt'):
            file_number = int(filename.split('_')[1].split('.')[0])  # Extract the file number

            if file_number == 1:
                with open(os.path.join(allostery_folder_path_rep1, filename), 'r') as file:
                    column_A_data = [line.split()[0] for line in file.readlines()]
                    for i, value in enumerate(column_A_data, start=1):
                        ws1.cell(row=i + 5, column=1, value=value)
                        ws2.cell(row=i + 5, column=1, value=value)
                        if F_or_G == 'Factin':
                            ws3.cell(row=i + 5, column=1, value=value)
                            ws4.cell(row=i + 5, column=1, value=value)

            # Read the contents of column C from the .txt file
            with open(os.path.join(allostery_folder_path_rep1, filename), 'r') as file:
                column_C_data = [line.split()[2] for line in file.readlines()]

            if mutant in filename:
                # Calculate the column to paste data (every 4th column starting from target_column)
                paste_column_C = 3 + 6 * (file_number - 1)
                # Paste the data into the specified column
                for i, value in enumerate(column_C_data, start=1):
                    ws1.cell(row=i + 5, column=paste_column_C, value=value)
            if 'WT' in filename:
                # Calculate the column to paste data (every 4th column starting from target_column)
                paste_column_C = 6 + 6 * (file_number - allostery_filenumber_offset)
                # Paste the data into the specified column
                for i, value in enumerate(column_C_data, start=1):
                    ws1.cell(row=i + 5, column=paste_column_C, value=value)

    for filename in sorted(os.listdir(allostery_folder_path_rep2)):
        if filename.startswith('processed_') and filename.endswith('.txt'):
            file_number = int(filename.split('_')[1].split('.')[0])  # Extract the file number

            # Read the contents of column C from the .txt file
            with open(os.path.join(allostery_folder_path_rep2, filename), 'r') as file:
                column_C_data = [line.split()[2] for line in file.readlines()]

            if mutant in filename:
                # Calculate the column to paste data (every 4th column starting from target_column)
                paste_column_C = 4 + 6 * (file_number - 1)
                # Paste the data into the specified column
                for i, value in enumerate(column_C_data, start=1):
                    ws1.cell(row=i + 5, column=paste_column_C, value=value)
            if 'WT' in filename:
                # Calculate the column to paste data (every 4th column starting from target_column)
                paste_column_C = 7 + 6 * (file_number - allostery_filenumber_offset)
                # Paste the data into the specified column
                for i, value in enumerate(column_C_data, start=1):
                    ws1.cell(row=i + 5, column=paste_column_C, value=value)

    for filename in sorted(os.listdir(allostery_folder_path_rep3)):
        if filename.startswith('processed_') and filename.endswith('.txt'):
            file_number = int(filename.split('_')[1].split('.')[0])  # Extract the file number

            # Read the contents of column C from the .txt file
            with open(os.path.join(allostery_folder_path_rep3, filename), 'r') as file:
                column_C_data = [line.split()[2] for line in file.readlines()]

            if mutant in filename:
                # Calculate the column to paste data (every 4th column starting from target_column)
                paste_column_C = 5 + 6 * (file_number - 1)
                # Paste the data into the specified column
                for i, value in enumerate(column_C_data, start=1):
                    ws1.cell(row=i + 5, column=paste_column_C, value=value)
            if 'WT' in filename:
                # Calculate the column to paste data (every 4th column starting from target_column)
                paste_column_C = 8 + 6 * (file_number - allostery_filenumber_offset)
                # Paste the data into the specified column
                for i, value in enumerate(column_C_data, start=1):
                    ws1.cell(row=i + 5, column=paste_column_C, value=value)

    
    if F_or_G == 'Factin':
        for filename in sorted(os.listdir(interprotomer_folder_path_rep1)):
            if filename.startswith('processed_') and filename.endswith('.txt'):
                file_number = int(filename.split('_')[1].split('.')[0])  # Extract the file number
    
                # Read the contents of column C from the .txt file
                with open(os.path.join(interprotomer_folder_path_rep1, filename), 'r') as file:
                    column_C_data = [line.split()[2] for line in file.readlines()]
    
                if mutant in filename:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column_C = 3 + 6 * (file_number - 1)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws4.cell(row=i + 5, column=paste_column_C, value=value)
                if 'WT' in filename:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column_C = 6 + 6 * (file_number - interprotomer_filenumber_offset)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws4.cell(row=i + 5, column=paste_column_C, value=value)
    
        for filename in sorted(os.listdir(interprotomer_folder_path_rep2)):
            if filename.startswith('processed_') and filename.endswith('.txt'):
                file_number = int(filename.split('_')[1].split('.')[0])  # Extract the file number
    
                # Read the contents of column C from the .txt file
                with open(os.path.join(interprotomer_folder_path_rep2, filename), 'r') as file:
                    column_C_data = [line.split()[2] for line in file.readlines()]
    
                if mutant in filename:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column_C = 4 + 6 * (file_number - 1)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws4.cell(row=i + 5, column=paste_column_C, value=value)
                if 'WT' in filename:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column_C = 7 + 6 * (file_number - interprotomer_filenumber_offset)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws4.cell(row=i + 5, column=paste_column_C, value=value)
    
        for filename in sorted(os.listdir(interprotomer_folder_path_rep3)):
            if filename.startswith('processed_') and filename.endswith('.txt'):
                file_number = int(filename.split('_')[1].split('.')[0])  # Extract the file number
    
                # Read the contents of column C from the .txt file
                with open(os.path.join(interprotomer_folder_path_rep3, filename), 'r') as file:
                    column_C_data = [line.split()[2] for line in file.readlines()]
    
                if mutant in filename:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column_C = 5 + 6 * (file_number - 1)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws4.cell(row=i + 5, column=paste_column_C, value=value)
                if 'WT' in filename:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column_C = 8 + 6 * (file_number - interprotomer_filenumber_offset)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws4.cell(row=i + 5, column=paste_column_C, value=value)

    if F_or_G == 'Factin':
        for filename in sorted(os.listdir(interstrand_folder_path_rep1)):
            if filename.startswith('processed_') and filename.endswith('.txt'):
                file_number = int(filename.split('_')[1].split('.')[0])  # Extract the file number

                # Read the contents of column C from the .txt file
                with open(os.path.join(interstrand_folder_path_rep1, filename), 'r') as file:
                    column_C_data = [line.split()[2] for line in file.readlines()]

                if mutant in filename:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column_C = 3 + 6 * (file_number - 1)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws3.cell(row=i + 5, column=paste_column_C, value=value)
                if 'WT' in filename:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column_C = 6 + 6 * (file_number - interstrand_filenumber_offset)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws3.cell(row=i + 5, column=paste_column_C, value=value)

        for filename in sorted(os.listdir(interstrand_folder_path_rep2)):
            if filename.startswith('processed_') and filename.endswith('.txt'):
                file_number = int(filename.split('_')[1].split('.')[0])  # Extract the file number

                # Read the contents of column C from the .txt file
                with open(os.path.join(interstrand_folder_path_rep2, filename), 'r') as file:
                    column_C_data = [line.split()[2] for line in file.readlines()]

                if mutant in filename:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column_C = 4 + 6 * (file_number - 1)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws3.cell(row=i + 5, column=paste_column_C, value=value)
                if 'WT' in filename:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column_C = 7 + 6 * (file_number - interstrand_filenumber_offset)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws3.cell(row=i + 5, column=paste_column_C, value=value)

        for filename in sorted(os.listdir(interstrand_folder_path_rep3)):
            if filename.startswith('processed_') and filename.endswith('.txt'):
                file_number = int(filename.split('_')[1].split('.')[0])  # Extract the file number

                # Read the contents of column C from the .txt file
                with open(os.path.join(interstrand_folder_path_rep3, filename), 'r') as file:
                    column_C_data = [line.split()[2] for line in file.readlines()]

                if mutant in filename:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column_C = 5 + 6 * (file_number - 1)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws3.cell(row=i + 5, column=paste_column_C, value=value)
                if 'WT' in filename:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column_C = 8 + 6 * (file_number - interstrand_filenumber_offset)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws3.cell(row=i + 5, column=paste_column_C, value=value)


    for filename in sorted(os.listdir(nucleotide_folder_path_rep1)):
        if filename.startswith('processed_') and filename.endswith('.txt'):
            file_number = int(filename.split('_')[1].split('.')[0])  # Extract the file number

            # Read the contents of column C from the .txt file
            with open(os.path.join(nucleotide_folder_path_rep1, filename), 'r') as file:
                column_C_data = [line.split()[2] for line in file.readlines()]

            if mutant in filename:
                # Calculate the column to paste data (every 4th column starting from target_column)
                paste_column_C = 3 + 6 * (file_number - 1)
                # Paste the data into the specified column
                for i, value in enumerate(column_C_data, start=1):
                    if F_or_G == 'Factin':
                        ws2.cell(row=i + 5, column=paste_column_C, value=value)
                    if F_or_G == 'Gactin':
                        ws2.cell(row=i + 5, column=paste_column_C, value=value)
            if 'WT' in filename:
                # Calculate the column to paste data (every 4th column starting from target_column)
                paste_column_C = 6 + 6 * (file_number - nucleotide_filenumber_offset)
                # Paste the data into the specified column
                for i, value in enumerate(column_C_data, start=1):
                    if F_or_G == 'Factin':
                        ws2.cell(row=i + 5, column=paste_column_C, value=value)
                    if F_or_G == 'Gactin':
                        ws2.cell(row=i + 5, column=paste_column_C, value=value)

    for filename in sorted(os.listdir(nucleotide_folder_path_rep2)):
        if filename.startswith('processed_') and filename.endswith('.txt'):
            file_number = int(filename.split('_')[1].split('.')[0])  # Extract the file number

            # Read the contents of column C from the .txt file
            with open(os.path.join(nucleotide_folder_path_rep2, filename), 'r') as file:
                column_C_data = [line.split()[2] for line in file.readlines()]

            if mutant in filename:
                # Calculate the column to paste data (every 4th column starting from target_column)
                paste_column_C = 4 + 6 * (file_number - 1)
                # Paste the data into the specified column
                for i, value in enumerate(column_C_data, start=1):
                    if F_or_G == 'Factin':
                        ws2.cell(row=i + 5, column=paste_column_C, value=value)
                    if F_or_G == 'Gactin':
                        ws2.cell(row=i + 5, column=paste_column_C, value=value)
            if 'WT' in filename:
                # Calculate the column to paste data (every 4th column starting from target_column)
                paste_column_C = 7 + 6 * (file_number - nucleotide_filenumber_offset)
                # Paste the data into the specified column
                for i, value in enumerate(column_C_data, start=1):
                    if F_or_G == 'Factin':
                        ws2.cell(row=i + 5, column=paste_column_C, value=value)
                    if F_or_G == 'Gactin':
                        ws2.cell(row=i + 5, column=paste_column_C, value=value)

    for filename in sorted(os.listdir(nucleotide_folder_path_rep3)):
        if filename.startswith('processed_') and filename.endswith('.txt'):
            file_number = int(filename.split('_')[1].split('.')[0])  # Extract the file number

            # Read the contents of column C from the .txt file
            with open(os.path.join(nucleotide_folder_path_rep3, filename), 'r') as file:
                column_C_data = [line.split()[2] for line in file.readlines()]

            if mutant in filename:
                # Calculate the column to paste data (every 4th column starting from target_column)
                paste_column_C = 5 + 6 * (file_number - 1)
                # Paste the data into the specified column
                for i, value in enumerate(column_C_data, start=1):
                    if F_or_G == 'Factin':
                        ws2.cell(row=i + 5, column=paste_column_C, value=value)
                    if F_or_G == 'Gactin':
                        ws2.cell(row=i + 5, column=paste_column_C, value=value)
            if 'WT' in filename:
                # Calculate the column to paste data (every 4th column starting from target_column)
                paste_column_C = 8 + 6 * (file_number - nucleotide_filenumber_offset)
                # Paste the data into the specified column
                for i, value in enumerate(column_C_data, start=1):
                    if F_or_G == 'Factin':
                        ws2.cell(row=i + 5, column=paste_column_C, value=value)
                    if F_or_G == 'Gactin':
                        ws2.cell(row=i + 5, column=paste_column_C, value=value)
    
    
    wb.save(f'{path_for_compiled_data}{mutant}_{F_or_G}_DISTANCES.xlsx')


def process_rmsdeviation(rmsdev_path, F_or_G, mutant, target_chain):
    delimiter = ' '
    file_counter = 1

    naming_phrases = [
        'C-alpha.',
        'C-alpha_EQ.',
        'Protein.',
        'Protein_EQ.'
    ]

    for naming_phrase in naming_phrases:
        for filename in os.listdir(rmsdev_path):
            if f'{mutant}_Chain_{target_chain}_rmsdev' in filename and '.xvg' in filename and naming_phrase in filename:
                input_file = os.path.join(rmsdev_path, filename)
                output_file = os.path.join(rmsdev_path, f"processed_{file_counter}_{mutant}_{F_or_G}_{naming_phrase}txt")

                with open(input_file, 'r') as f:
                    lines = f.readlines()
                # Remove lines starting with '#' or '@' (common in .xvg files)
                lines = [line.strip() for line in lines if
                         not line.startswith('#') and not line.startswith('@') and line.strip()]
                lines = [line.strip() for line in lines if
                         not line.startswith('@') and not line.startswith('@') and line.strip()]
                lines = [line.strip() for line in lines if
                         not line.startswith('&') and not line.startswith('&') and line.strip()]
                # Split the lines based on the delimiter and create a list of tuples with two columns
                two_columns = []
                for line in lines:
                    split_line = line.split(delimiter)
                    two_columns.append((split_line[0], ' '.join(split_line[1:])))
                # Calculate the formula (B * 10) and write it to the corresponding cell in column C
                with open(output_file, 'w') as f_out:
                    for idx, (col_A, col_B) in enumerate(two_columns):
                        try:
                            # Convert col_B to float and multiply by 10 if it's not zero
                            value_b = float(col_B)
                            if value_b != 0:
                                value_c = value_b * 10
                            else:
                                value_c = 0  # Keep value as 0 if col_B is zero
                        except ValueError:
                            value_c = 'Invalid'  # Or handle the invalid value case differently
                        # Write the values to the output file (or you can write to Excel using a library like openpyxl)
                        f_out.write(f"{col_A}\t{col_B}\t{value_c}\n")

                file_counter += 1

    for naming_phrase in naming_phrases:
        for filename in os.listdir(rmsdev_path):
            if f'WT_Chain_{target_chain}_rmsdev' in filename and '.xvg' in filename and naming_phrase in filename:
                input_file = os.path.join(rmsdev_path, filename)
                output_file = os.path.join(rmsdev_path, f"processed_{file_counter}_WT_{F_or_G}_{naming_phrase}txt")

                with open(input_file, 'r') as f:
                    lines = f.readlines()
                # Remove lines starting with '#' or '@' (common in .xvg files)
                lines = [line.strip() for line in lines if
                         not line.startswith('#') and not line.startswith('@') and line.strip()]
                lines = [line.strip() for line in lines if
                         not line.startswith('@') and not line.startswith('@') and line.strip()]
                lines = [line.strip() for line in lines if
                         not line.startswith('&') and not line.startswith('&') and line.strip()]
                # Split the lines based on the delimiter and create a list of tuples with two columns
                two_columns = []
                for line in lines:
                    split_line = line.split(delimiter)
                    two_columns.append((split_line[0], ' '.join(split_line[1:])))
                # Calculate the formula (B * 10) and write it to the corresponding cell in column C
                with open(output_file, 'w') as f_out:
                    for idx, (col_A, col_B) in enumerate(two_columns):
                        try:
                            # Convert col_B to float and multiply by 10 if it's not zero
                            value_b = float(col_B)
                            if value_b != 0:
                                value_c = value_b * 10
                            else:
                                value_c = 0  # Keep value as 0 if col_B is zero
                        except ValueError:
                            value_c = 'Invalid'  # Or handle the invalid value case differently
                        # Write the values to the output file (or you can write to Excel using a library like openpyxl)
                        f_out.write(f"{col_A}\t{col_B}\t{value_c}\n")

                file_counter += 1


def compile_rmsdeviation(rmsdev_path1, rmsdev_path2, rmsdev_path3, F_or_G, mutant, number_of_residues, path_for_compiled_data, target_chain_number):
    # Load the workbook
    wb = load_workbook(f'{path_for_compiled_data}{mutant}_{F_or_G}_COMPILED.xlsx')
    # Create a new worksheet named 'RMSF'
    ws8 = wb.create_sheet(title='RMSDEVIATION')
    ws8.merge_cells('C1:N1')
    ws8['C1'] = f'{mutant}'
    ws8.merge_cells('O1:Z1')
    ws8['O1'] = 'WT'

    ws8.merge_cells('C2:E2')
    ws8['C2'] = 'C-alpha'
    ws8.merge_cells('F2:H2')
    ws8['F2'] = 'C-alpha EQ'
    ws8.merge_cells('I2:K2')
    ws8['I2'] = 'Protein'
    ws8.merge_cells('L2:N2')
    ws8['L2'] = 'Protein EQ'
    ws8['B3'] = 'Residue Number'
    ws8['C3'] = 'Replicate1'
    ws8['D3'] = 'Replicate2'
    ws8['E3'] = 'Replicate3'
    ws8['F3'] = 'Replicate1'
    ws8['G3'] = 'Replicate2'
    ws8['H3'] = 'Replicate3'
    ws8['I3'] = 'Replicate1'
    ws8['J3'] = 'Replicate2'
    ws8['K3'] = 'Replicate3'
    ws8['L3'] = 'Replicate1'
    ws8['M3'] = 'Replicate2'
    ws8['N3'] = 'Replicate3'

    ws8.merge_cells('O2:Q2')
    ws8['O2'] = 'C-alpha'
    ws8.merge_cells('R2:T2')
    ws8['R2'] = 'C-alpha EQ'
    ws8.merge_cells('U2:W2')
    ws8['U2'] = 'Protein'
    ws8.merge_cells('X2:Z2')
    ws8['X2'] = 'Protein EQ'
    ws8['O3'] = 'Replicate1'
    ws8['P3'] = 'Replicate2'
    ws8['Q3'] = 'Replicate3'
    ws8['R3'] = 'Replicate1'
    ws8['S3'] = 'Replicate2'
    ws8['T3'] = 'Replicate3'
    ws8['U3'] = 'Replicate1'
    ws8['V3'] = 'Replicate2'
    ws8['W3'] = 'Replicate3'
    ws8['X3'] = 'Replicate1'
    ws8['Y3'] = 'Replicate2'
    ws8['Z3'] = 'Replicate3'


    # Copy headers to other columns for individual chains
    source_range = ws8['B1:Z3']
    copied_data = []
    for row in source_range:
        copied_data.append([cell.value for cell in row])

    # Paste the copied values to the destination range (AB1)
    for row_idx, row in enumerate(copied_data, start=1):
        for col_idx, value in enumerate(row, start=1):
            col_value = 28
            ws8.cell(row=row_idx, column=col_idx + col_value).value = value

    input_text_file = f'{rmsdev_path1}processed_1_{mutant}_{F_or_G}_C-alpha.txt'
    output_excel_file = f'{rmsdev_path1}processed_1_{mutant}_{F_or_G}_C-alpha.xlsx'
    # Read the text file into a pandas DataFrame
    data = pd.read_csv(input_text_file, sep='\t')  # Modify 'sep' based on your text file's delimiter
    # Write the DataFrame to an Excel file
    data.to_excel(output_excel_file, index=False)
    source_wb = load_workbook(output_excel_file)
    source_ws = source_wb.active
    # copy residue number from data file
    for row_idx, row in enumerate(source_ws[f'A1:A{number_of_residues + 1}'], start=1):
        for col_idx, cell in enumerate(row, start=1):
            dest_cell = ws8.cell(row=row_idx + int('B4'[1:]) - 1,
                                            column=coordinate_to_tuple('B4')[1])
            dest_cell.value = cell.value

    # for each processed txt file, isolate file number and copy column C to corresponding column
    for filename in os.listdir(rmsdev_path1):
        if mutant in filename:
            start_col = 3
            file_offset = 1
        if 'WT' in filename:
            start_col = 15
            file_offset = 5
        if filename.startswith("processed_") and filename.endswith('.txt'):
            file_path = os.path.join(rmsdev_path1, filename)
            file_number = int(filename.split('_')[1])
            delimiter = '\t'
            max_lines = number_of_residues + 1
            with open(file_path, 'r') as txt_file:
                lines = txt_file.readlines()
                column_C_data = [line.split(delimiter)[2] for line in
                                 lines[:max_lines]]  # Assuming data is space-separated
            target_column = start_col + (file_number - file_offset) * 3
            for idx, data in enumerate(column_C_data, start=4):
                ws8.cell(row=idx, column=target_column).value = float(data) if data else None

    for filename in os.listdir(rmsdev_path2):
        if mutant in filename:
            start_col = 4
            file_offset = 1
        if 'WT' in filename:
            start_col = 16
            file_offset = 5
        if filename.startswith("processed_") and filename.endswith('.txt'):
            file_path = os.path.join(rmsdev_path2, filename)
            file_number = int(filename.split('_')[1])
            delimiter = '\t'
            max_lines = number_of_residues + 1
            with open(file_path, 'r') as txt_file:
                lines = txt_file.readlines()
                column_C_data = [line.split(delimiter)[2] for line in
                                 lines[:max_lines]]  # Assuming data is space-separated
            target_column = start_col + (file_number - file_offset) * 3
            for idx, data in enumerate(column_C_data, start=4):
                ws8.cell(row=idx, column=target_column).value = float(data) if data else None

    for filename in os.listdir(rmsdev_path3):
        if mutant in filename:
            start_col = 5
            file_offset = 1
        if 'WT' in filename:
            start_col = 17
            file_offset = 5
        if filename.startswith("processed_") and filename.endswith('.txt'):
            file_path = os.path.join(rmsdev_path3, filename)
            file_number = int(filename.split('_')[1])
            delimiter = '\t'
            max_lines = number_of_residues + 1
            with open(file_path, 'r') as txt_file:
                lines = txt_file.readlines()
                column_C_data = [line.split(delimiter)[2] for line in
                                 lines[:max_lines]]  # Assuming data is space-separated
            target_column = start_col + (file_number - file_offset) * 3
            for idx, data in enumerate(column_C_data, start=4):
                ws8.cell(row=idx, column=target_column).value = float(data) if data else None

    #if Factin, isolate chain C data
    if F_or_G == 'Factin':
        target_chain_start_row = ( target_chain_number * 375 ) + 4
        target_chain_end_row = target_chain_start_row + 374
        cell_range = {'AC4': f'B{target_chain_start_row}:Z{target_chain_end_row}'}
        for start_cell, source_range in cell_range.items():
            source_rows = ws8[source_range]
            target_cell = ws8[start_cell]

            for dest_row, source_row in enumerate(source_rows, start=1):
                for col_idx, cell in enumerate(source_row, start=1):
                    dest_cell = ws8.cell(row=target_cell.row + dest_row - 1, column=target_cell.column + col_idx - 1)
                    dest_cell.value = cell.value
    wb.save(f'{path_for_compiled_data}{mutant}_{F_or_G}_COMPILED.xlsx')


#for R312, did additional interprotomer distance analyses
def process_interprotomer_distances2(path_for_compiled_data, F_or_G, mutant, number_of_timepoints, interprotomer2_folder_path, chains, target_chain_number):
    
    if "Replicate1" in interprotomer2_folder_path and F_or_G == 'Factin':
        wb = load_workbook(f'{path_for_compiled_data}{mutant}_{F_or_G}_DISTANCES.xlsx')
        ws5 = wb.create_sheet(title='Interprotomer2')

    if "Replicate1" in interprotomer2_folder_path and F_or_G == 'Gactin':
        wb = load_workbook(f'{path_for_compiled_data}{mutant}_{F_or_G}_DISTANCES.xlsx')
        ws5 = wb.create_sheet(title='Interprotomer2')

    #max column is based on whatever sheet will need the most columns
    #here, interprotomer_naming is the longest. It's 20 rows * 3 replicates * 2 (mutant, WT) = 120
    #it's needed for determining where to stop merging header cells
    max_column = 120
    target_chain = chains[target_chain_number]
    col = 3
    file_counter = 1
    delimiter = ' '
    if F_or_G == 'Factin':
        col = 3
        file_counter = 1
        interprotomer2_chain1 = chains[target_chain_number - 2]
        interprotomer2_chain2 = chains[target_chain_number + 2]
        interprotomer2_naming = [
            f'Chain_{target_chain}_D244_Chain_{interprotomer2_chain1}_M325',
            f'Chain_{target_chain}_D244_Chain_{interprotomer2_chain1}_R290',
            f'Chain_{target_chain}_D244_Chain_{interprotomer2_chain1}_I287',
            f'Chain_{target_chain}_E205_Chain_{interprotomer2_chain1}_I287',
            f'Chain_{target_chain}_R62_Chain_{interprotomer2_chain1}_D288',
            f'Chain_{target_chain}_G63_Chain_{interprotomer2_chain1}_D286',
            f'Chain_{target_chain}_G63_Chain_{interprotomer2_chain1}_D288',
            f'Chain_{interprotomer2_chain2}_D244_Chain_{target_chain}_M325',
            f'Chain_{interprotomer2_chain2}_D244_Chain_{target_chain}_R290',
            f'Chain_{interprotomer2_chain2}_D244_Chain_{target_chain}_I287',
            f'Chain_{interprotomer2_chain2}_E205_Chain_{target_chain}_I287',
            f'Chain_{interprotomer2_chain2}_R62_Chain_{target_chain}_D288',
            f'Chain_{interprotomer2_chain2}_G63_Chain_{target_chain}_D286',
            f'Chain_{interprotomer2_chain2}_G63_Chain_{target_chain}_D288',
            f'Chain_{target_chain}_D292_Chain_{target_chain}_Y166',
            f'Chain_{target_chain}_I289_Chain_{target_chain}_Y166'
        ]
        # add headers to excel file
        if "Replicate1" in interprotomer2_folder_path:
            for interprotomer2_header in interprotomer2_naming:
                cell = ws5.cell(row=1, column=col)
                cell.value = interprotomer2_header
                col += 6

        # first, open each .xvg file, strip non-data lines and save it in new format
        # run for PBM files first
        for interprotomer2_name in interprotomer2_naming:
            for filename in os.listdir(interprotomer2_folder_path):
                if mutant in filename and interprotomer2_name in filename and filename.endswith('.xvg'):
                    input_file = os.path.join(interprotomer2_folder_path, filename)
                    output_file = os.path.join(interprotomer2_folder_path,
                                               f"{file_counter}_{mutant}_{F_or_G}_{interprotomer2_name}.txt")
                    with open(input_file, 'r') as f:
                        lines = f.readlines()
                    # Remove lines starting with '#' or '@' (common in .xvg files)
                    lines = [line.strip() for line in lines if
                             not line.startswith('#') and not line.startswith('@') and line.strip()]
                    two_columns = []
                    for line in lines:
                        split_line = line.split(delimiter)
                        two_columns.extend([(split_line[0], ' '.join(split_line[1:]))])
                    with open(output_file, 'w') as f_out:
                        for row in two_columns:
                            f_out.write('\t'.join(row) + '\n')
                    file_counter += 1
        # next run the same commands for WT files
        for interprotomer2_name in interprotomer2_naming:
            for filename in os.listdir(interprotomer2_folder_path):
                if 'WT' in filename and interprotomer2_name in filename and filename.endswith('.xvg'):
                    input_file = os.path.join(interprotomer2_folder_path, filename)
                    output_file = os.path.join(interprotomer2_folder_path,
                                               f"{file_counter}_WT_{F_or_G}_{interprotomer2_name}.txt")
                    with open(input_file, 'r') as f:
                        lines = f.readlines()
                    # Remove lines starting with '#' or '@' (common in .xvg files)
                    lines = [line.strip() for line in lines if
                             not line.startswith('#') and not line.startswith('@') and line.strip()]
                    two_columns = []
                    for line in lines:
                        split_line = line.split(delimiter)
                        two_columns.extend([(split_line[0], ' '.join(split_line[1:]))])
                    with open(output_file, 'w') as f_out:
                        for row in two_columns:
                            f_out.write('\t'.join(row) + '\n')
                    file_counter += 1

        for interprotomer2_name in interprotomer2_naming:
            for filename in os.listdir(interprotomer2_folder_path):
                if interprotomer2_name in filename and filename.endswith('.txt'):
                    input_file = os.path.join(interprotomer2_folder_path, filename)
                    output_file = os.path.join(interprotomer2_folder_path, f"processed_{filename}")
                    # Read the contents of the .txt file
                    with open(input_file, 'r') as txt_file:
                        lines = txt_file.readlines()

                    # Apply formula to multiply values by 10 (if the value is not 0)
                    modified_lines = []
                    for line in lines:
                        line_data = line.strip().split('\t')  # Assuming tab-separated data, modify delimiter if needed
                        if len(line_data) >= 2:
                            try:
                                value = float(line_data[1])
                                if value != 0:
                                    line_data.append(str(value * 10))
                                else:
                                    line_data.append('0')
                            except ValueError:
                                line_data.append('Invalid')  # Or handle the invalid value case differently

                        modified_lines.append('\t'.join(line_data))

                    # Write the modified data to a new output file
                    with open(output_file, 'w') as output_txt_file:
                        output_txt_file.write('\n'.join(modified_lines))
    
    if "Replicate1" in interprotomer2_folder_path:
        # merge header cells
        for col in range(3, max_column, 6):
            if F_or_G == 'Factin':
                merge_range = f'{ws5.cell(row=1, column=col).coordinate}:{ws5.cell(row=1, column=col + 5).coordinate}'
                ws5.merge_cells(merge_range)

        # Assign alternating values for PBM and "WT" to individual cells
        for col in range(3, max_column, 6):
            if F_or_G == 'Factin':
                cell = ws5.cell(row=2, column=col)
                cell.value = f'{mutant}'
        for col in range(6, max_column, 6):
            if F_or_G == 'Factin':
                cell = ws5.cell(row=2, column=col)
                cell.value = 'WT'

        # merge PBM and WT header cells
        for col in range(3, max_column, 6):
            if F_or_G == 'Factin':
                merge_range = f'{ws5.cell(row=2, column=col).coordinate}:{ws5.cell(row=2, column=col + 2).coordinate}'
                ws5.merge_cells(merge_range)

        # add Replicate headers
        for col in range(3, max_column + 1, 3):
            if F_or_G == 'Factin':
                cell = ws5.cell(row=5, column=col)
                cell.value = "Replicate1"
            col += 1
            if F_or_G == 'Factin':
                cell = ws5.cell(row=5, column=col)
                cell.value = "Replicate2"
            col += 1
            if F_or_G == 'Factin':
                cell = ws5.cell(row=5, column=col)
                cell.value = "Replicate3"
            col += 1

        # add formula for converting time to ns
        for row in range(6, 7 + number_of_timepoints):
            formula = f'=A{row}/1000'  # Change the formula as needed
            if F_or_G == 'Factin':
                ws5[f'B{row}'].value = formula

        # add headers for formulas
        if F_or_G == 'Factin':
            ws5['B3'] = 'Average Distance'
            ws5['B4'] = 'Occupancy %'

        # add formulas to rows 3/4 to calculate average distance, and determine occupancy percentage
        for col in range(3, max_column + 1, 3):
            col_letter = get_column_letter(col)
            next_col_letter = get_column_letter(col + 2)

            formula1 = f'=AVERAGE({col_letter}6:{next_col_letter}{number_of_timepoints + 6})'
            formula2 = f'=(COUNTIF({col_letter}6:{next_col_letter}{number_of_timepoints + 6}, "<=3.5") / COUNT({col_letter}6:{next_col_letter}{number_of_timepoints + 6}))'
            if F_or_G == 'Factin':
                ws5[f"{col_letter}3"].value = formula1
                ws5[f"{col_letter}4"].value = formula2

        wb.save(f'{path_for_compiled_data}{mutant}_{F_or_G}_DISTANCES.xlsx')

#for R312, did additional interprotomer distance analyses
def compile_interprotomer_distances2(F_or_G, mutant, path_for_compiled_data,
                      interprotomer2_folder_path_rep1, interprotomer2_folder_path_rep2, interprotomer2_folder_path_rep3):
    if F_or_G == 'Factin':
        # offset number is the start number of the WT files, used for calculating what column to paste into
        interprotomer2_filenumber_offset = 17


    wb = load_workbook(f'{path_for_compiled_data}{mutant}_{F_or_G}_DISTANCES.xlsx')
    if F_or_G == 'Factin':
        ws5 = wb['Interprotomer2']

    if F_or_G == 'Factin':
        for filename in sorted(os.listdir(interprotomer2_folder_path_rep1)):
            if filename.startswith('processed_') and filename.endswith('.txt'):
                file_number = int(filename.split('_')[1].split('.')[0])  # Extract the file number

                # Read the contents of column C from the .txt file
                with open(os.path.join(interprotomer2_folder_path_rep1, filename), 'r') as file:
                    column_C_data = [line.split()[2] for line in file.readlines()]

                if mutant in filename:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column_C = 3 + 6 * (file_number - 1)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws5.cell(row=i + 5, column=paste_column_C, value=value)
                if 'WT' in filename:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column_C = 6 + 6 * (file_number - interprotomer2_filenumber_offset)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws5.cell(row=i + 5, column=paste_column_C, value=value)

        for filename in sorted(os.listdir(interprotomer2_folder_path_rep2)):
            if filename.startswith('processed_') and filename.endswith('.txt'):
                file_number = int(filename.split('_')[1].split('.')[0])  # Extract the file number

                # Read the contents of column C from the .txt file
                with open(os.path.join(interprotomer2_folder_path_rep2, filename), 'r') as file:
                    column_C_data = [line.split()[2] for line in file.readlines()]

                if mutant in filename:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column_C = 4 + 6 * (file_number - 1)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws5.cell(row=i + 5, column=paste_column_C, value=value)
                if 'WT' in filename:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column_C = 7 + 6 * (file_number - interprotomer2_filenumber_offset)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws5.cell(row=i + 5, column=paste_column_C, value=value)

        for filename in sorted(os.listdir(interprotomer2_folder_path_rep3)):
            if filename.startswith('processed_') and filename.endswith('.txt'):
                file_number = int(filename.split('_')[1].split('.')[0])  # Extract the file number

                # Read the contents of column C from the .txt file
                with open(os.path.join(interprotomer2_folder_path_rep3, filename), 'r') as file:
                    column_C_data = [line.split()[2] for line in file.readlines()]

                if mutant in filename:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column_C = 5 + 6 * (file_number - 1)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws5.cell(row=i + 5, column=paste_column_C, value=value)
                if 'WT' in filename:
                    # Calculate the column to paste data (every 4th column starting from target_column)
                    paste_column_C = 8 + 6 * (file_number - interprotomer2_filenumber_offset)
                    # Paste the data into the specified column
                    for i, value in enumerate(column_C_data, start=1):
                        ws5.cell(row=i + 5, column=paste_column_C, value=value)

    wb.save(f'{path_for_compiled_data}{mutant}_{F_or_G}_DISTANCES.xlsx')




